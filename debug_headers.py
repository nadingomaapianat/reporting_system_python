#!/usr/bin/env python3

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from dynamic_report_endpoints import generate_excel_report
from openpyxl import load_workbook

# Test the Excel generation function directly
def test_excel_headers():
    print("Testing Excel generation with header config...")
    
    # Test data
    columns = ["id", "title", "format"]
    data_rows = [
        ["1", "Test Report 1", "excel"],
        ["2", "Test Report 2", "word"]
    ]
    
    # Test header configuration
    header_config = {
        "includeHeader": True,
        "title": "CUSTOM TEST TITLE",
        "subtitle": "CUSTOM TEST SUBTITLE",
        "fontColor": "#FF0000"
    }
    
    print(f"Header config: {header_config}")
    
    try:
        # Generate the Excel file
        result = generate_excel_report(columns, data_rows, header_config)
        print(f"✅ Excel generation successful!")
        
        # The result is a FileResponse, so we need to get the file path
        # Let's check if the file was created in the exports directory
        import glob
        export_files = glob.glob("exports/dynamic_report_*.xlsx")
        if export_files:
            filename = max(export_files, key=os.path.getctime)  # Get the most recent file
            print(f"✅ Found generated file: {filename}")
        else:
            print("❌ No export files found")
            return
        
        # Load and check the Excel file
        wb = load_workbook(filename)
        ws = wb.active
        
        print(f"✅ Excel file loaded successfully")
        print(f"Sheet title: {ws.title}")
        print(f"Max row: {ws.max_row}")
        print(f"Max column: {ws.max_column}")
        
        # Check all rows
        for row in range(1, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"Row {row}: {row_data}")
        
        # Check if the column headers are styled
        print(f"\nChecking column headers at row 6:")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=6, column=col)
            print(f"  Column {col}: '{cell.value}' - Font: {cell.font.bold if cell.font else 'None'}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_excel_headers()
