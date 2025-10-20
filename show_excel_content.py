#!/usr/bin/env python3

import os
from openpyxl import load_workbook

def show_excel_content():
    print("🔍 Checking Excel file content...")
    
    # Find the most recent export file
    export_files = [f for f in os.listdir('exports') if f.startswith('dynamic_report_') and f.endswith('.xlsx')]
    if not export_files:
        print("❌ No export files found")
        return
    
    # Get the most recent file
    latest_file = max(export_files, key=lambda f: os.path.getctime(os.path.join('exports', f)))
    filepath = os.path.join('exports', latest_file)
    
    print(f"📁 Latest file: {latest_file}")
    print(f"📅 Created: {os.path.getctime(filepath)}")
    
    # Load the Excel file
    wb = load_workbook(filepath)
    ws = wb.active
    
    print(f"📊 Sheet info: {ws.max_row} rows, {ws.max_column} columns")
    print()
    
    # Show all rows with their content and styling
    print("📋 COMPLETE EXCEL CONTENT:")
    print("=" * 50)
    
    for row in range(1, ws.max_row + 1):
        row_data = []
        row_styles = []
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            font_bold = cell.font.bold if cell.font else False
            font_size = cell.font.size if cell.font else None
            font_color = cell.font.color.rgb if cell.font and cell.font.color else None
            fill_color = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
            
            row_data.append(str(value) if value is not None else "")
            row_styles.append(f"Bold:{font_bold},Size:{font_size},Color:{font_color},Fill:{fill_color}")
        
        print(f"Row {row:2d}: {row_data}")
        if any(style != "Bold:False,Size:11.0,Color:Values must be of type <class 'str'>,Fill:00000000" for style in row_styles):
            print(f"      Styles: {row_styles}")
        print()
    
    print("=" * 50)
    print("✅ This shows ALL content in the Excel file")
    print("📌 Custom headers are in the FIRST FEW ROWS")
    print("📌 Data table starts from the row with column headers")

if __name__ == "__main__":
    show_excel_content()
