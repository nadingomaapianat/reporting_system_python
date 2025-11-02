import asyncio
import aiohttp
from typing import Dict, Any, Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import io
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')

class ExcelService:
    def __init__(self):
        self.api_service = None
        self.database_service = None

    def set_services(self, api_service, database_service):
        self.api_service = api_service
        self.database_service = database_service

    async def generate_excel_report(self, report_data: Dict[str, Any], report_config: Dict[str, Any]) -> bytes:
        """Generate Excel report from data and configuration using reusable functions"""
        try:
            from routes.route_utils import generate_excel_report
            
            # For now, generate a basic report
            data_rows = [['Controls Dashboard Report', 'Generated Successfully']]
            columns = ['Title', 'Status']
            
            return generate_excel_report(columns, data_rows, report_config)
            
        except Exception as e:
            print(f"Error generating Excel report: {e}")
            raise e

    async def generate_controls_excel(self, controls_data: Dict[str, Any], startDate: str, endDate: str, header_config: Dict[str, Any], cardType: str = None, onlyCard: bool = False, onlyOverallTable: bool = False, onlyChart: bool = False) -> bytes:
        """Generate Excel report for controls dashboard - matches PDF format exactly"""
        try:
            from routes.route_utils import generate_excel_report
            import re
            
            # Get the data from controls_data
            data = controls_data.get(cardType, []) if controls_data else []
            from routes.route_utils import write_debug
            write_debug(f"Excel export - cardType={cardType}, data type={type(data)}, len={len(data) if isinstance(data, list) else 'N/A'}")
            
            columns = []
            data_rows = []
            
            # Define column name formatter (same as PDF)
            def format_column_name(key):
                """Convert snake_case or camelCase to Title Case"""
                formatted = re.sub(r'[_]|([a-z])([A-Z])', r'\1 \2', str(key))
                return formatted.title()
            
            write_debug(f"Generating controls Excel report for ,,,,,,,,, {startDate} to {endDate}")
            
            # CHART EXPORT - same logic as PDF
            if onlyChart and cardType:
                if isinstance(data, list) and data:
                    first_item = data[0] if data else {}
                    if isinstance(first_item, dict):
                        keys = list(first_item.keys())
                        if len(keys) >= 2:
                            key1 = keys[0]
                            key2 = keys[-1]
                            label1 = format_column_name(key1)
                            label2 = format_column_name(key2)
                            columns = [label1, label2]
                            
                            for item in data:
                                name = item.get(key1, "N/A")
                                value = item.get(key2, 0)
                                data_rows.append([name, str(value)])
                        else:
                            key1 = keys[0]
                            columns = [format_column_name(key1), "Value"]
                            for item in data:
                                name = item.get(key1, "N/A")
                                data_rows.append([name, 0])
                else:
                    data_rows = [["No data available", "0"]]
                    columns = ["Label", "Value"]
                
                # Add chart type configuration (same as PDF)
                default_type_by_card = {
                    "quarterlyControlCreationTrend": "line",
                    "controlsByType": "pie",
                    "antiFraudDistribution": "pie",
                    "controlsPerLevel": "bar",
                    "controlExecutionFrequency": "bar",
                    "numberOfControlsPerComponent": "bar",
                    "departmentDistribution": "bar",
                    "numberOfControlsByIcofrStatus": "pie",
                }
                
                # Priority: renderType/chartType from query > cardType default > "bar"
                chart_type = header_config.get("chartType") or header_config.get("chart_type")
                if chart_type and chart_type in {"bar", "line", "pie"}:
                    write_debug(f"Using provided chart type: {chart_type}")
                else:
                    chart_type = default_type_by_card.get(cardType, "bar")
                    write_debug(f"Using default chart type for {cardType}: {chart_type}")
                
                header_config['chart_type'] = chart_type
                
                # Extract chart data for visualization (labels and values from data_rows)
                chart_labels = []
                chart_values = []
                for row in data_rows:
                    if len(row) >= 2:
                        chart_labels.append(str(row[0]))  # First column (label)
                        try:
                            chart_values.append(float(row[1]))  # Second column (value)
                        except:
                            chart_values.append(0)
                
                # Add chart data to header_config
                if len(chart_labels) > 0 and len(chart_values) > 0:
                    header_config['chart_data'] = {
                        'labels': chart_labels,
                        'values': chart_values
                    }
                    write_debug(f"Chart export: Prepared chart_data with {len(chart_labels)} items, type={chart_type}")
                
                return generate_excel_report(columns, data_rows, header_config)
            
            # TABLE EXPORT - same logic as PDF (dynamic column extraction)
            elif onlyOverallTable and cardType:
                # Generic table builder: derive columns from keys of first row and add index column
                if isinstance(data, list) and len(data) > 0:
                    first_item = data[0]
                    if isinstance(first_item, dict):
                        raw_keys = list(first_item.keys())
                        columns = ['#'] + [format_column_name(k) for k in raw_keys]
                        data_rows = []
                        for i, row in enumerate(data, 1):
                            if isinstance(row, dict):
                                values = [str(row.get(k, '')) for k in raw_keys]
                                data_rows.append([str(i)] + values)
                            elif isinstance(row, (list, tuple)):
                                vals = [str(v) for v in row]
                                data_rows.append([str(i)] + vals)
                            else:
                                data_rows.append([str(i), str(row)])
                    elif isinstance(first_item, (list, tuple)):
                        num_cols = len(first_item)
                        columns = ['#'] + [f'C{idx+1}' for idx in range(num_cols)]
                        data_rows = []
                        for i, row in enumerate(data, 1):
                            vals = [str(v) for v in (row if isinstance(row, (list, tuple)) else [row])]
                            data_rows.append([str(i)] + vals)
                    else:
                        columns = ['#', 'Value']
                        data_rows = [[str(i+1), str(v)] for i, v in enumerate(data)]
                else:
                    columns = ['#', 'Value']
                    data_rows = [['1', 'No data available']]
                
                write_debug(f"About to call generate_excel_report with chart_data={header_config.get('chart_data') is not None}")
                write_debug(f"header_config keys: {list(header_config.keys())}")
                result = generate_excel_report(columns, data_rows, header_config)
                write_debug(f"Excel report generated, returning {len(result) if result else 0} bytes")
                return result
            
            # CARD SUMMARY EXPORT - same logic as PDF
            elif onlyCard and cardType:
                write_debug(f"Generating risks Excel report for mmmmm {start_date} to {end_date}")
                write_debug(f"card_type={card_type}, only_card={only_card}, only_overall_table={only_overall_table}, only_chart={only_chart}")
               
                # Handle both dict and list data
                if isinstance(data, list) and data:
                    first_item = data[0] if data else {}
                    
                    if isinstance(first_item, str):
                        columns = ["#", "Control Name"]
                        data_rows = []
                        for i, item in enumerate(data, 1):
                            data_rows.append([str(i), str(item)])
                    elif isinstance(first_item, dict):
                        if cardType in ['pendingPreparer', 'pendingChecker', 'pendingReviewer', 'pendingAcceptance', 'testsPendingPreparer', 'testsPendingChecker', 'testsPendingReviewer', 'testsPendingAcceptance','unmappedControls','unmappedIcofrControls','unmappedNonIcofrControls','totalControls']:
                            columns = ["#", "Control Code", "Control Name"]
                            data_rows = []
                            for i, item in enumerate(data, 1):
                                if isinstance(item, dict):
                                    data_rows.append([
                                        str(i),
                                        item.get('control_code', item.get('code', 'N/A')),
                                        item.get('control_name', item.get('name', 'N/A'))
                                    ])
                                else:
                                    data_rows.append([str(i), 'N/A', 'N/A'])
                        else:
                            columns = ["#", "Code", "Name"]
                            data_rows = []
                            for i, item in enumerate(data, 1):
                                if isinstance(item, dict):
                                    data_rows.append([
                                        str(i),
                                        item.get('code', 'N/A'),
                                        item.get('name', 'N/A')
                                    ])
                                else:
                                    data_rows.append([str(i), 'N/A', 'N/A'])
                elif isinstance(data, dict):
                    columns = ["Metric", "Value"]
                    data_rows = [[key, str(value)] for key, value in data.items()]
                else:
                    columns = ["Metric", "Value"]
                    data_rows = [["No data available", "N/A"]]
                
                return generate_excel_report(columns, data_rows, header_config)
            
            else:
                # Generate basic report for other cases
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = header_config.get('title', 'Controls Report')
                ws['A1'] = 'Controls Dashboard Report'
                ws['A2'] = 'Generated Successfully'
                
                from io import BytesIO
                output = BytesIO()
                wb.save(output)
                return output.getvalue()
        except Exception as e:
            print(f"Error generating controls Excel: {e}")
            raise e
    
    async def generate_incidents_excel(self, incidents_data: Dict[str, Any], start_date: str, end_date: str, header_config: Dict[str, Any], card_type: str = None, only_card: bool = False, only_overall_table: bool = False, only_chart: bool = False) -> bytes:
        """Generate incidents Excel report mirroring risks/controls behavior."""
        try:
            from routes.route_utils import generate_excel_report, write_debug
            import re
            write_debug(f"Generating incidents Excel report for {start_date} to {end_date}")
            write_debug(f"card_type={card_type}, only_card={only_card}, only_overall_table={only_overall_table}, only_chart={only_chart}")

            data = incidents_data.get(card_type) if card_type else incidents_data.get('list')
            data = data or []
            write_debug(f"Incidents Excel export - card_type={card_type}, data type={type(data)}, len={len(data) if isinstance(data, list) else 'N/A'}")

            columns: List[str] = []
            data_rows: List[List[str]] = []

            def format_column_name(key: str) -> str:
                return re.sub(r'[_]|([a-z])([A-Z])', r'\1 \2', str(key)).title()

            # CHART EXPORT
            if only_chart and card_type:
                if isinstance(data, list) and data:
                    first_item = data[0]
                    if isinstance(first_item, dict):
                        keys = list(first_item.keys())
                        if len(keys) >= 2:
                            key1 = keys[0]
                            key2 = keys[-1]
                            columns = [format_column_name(key1), format_column_name(key2)]
                            for item in data:
                                name = item.get(key1, "N/A")
                                value = item.get(key2, 0)
                                data_rows.append([name, str(value)])
                        else:
                            key1 = keys[0]
                            columns = [format_column_name(key1), "Value"]
                            for item in data:
                                name = item.get(key1, "N/A")
                                data_rows.append([name, 0])
                    else:
                        data_rows = [[str(first_item), "0"]]
                        columns = ["Label", "Value"]
                else:
                    data_rows = [["No data available", "0"]]
                    columns = ["Label", "Value"]

                default_type_by_card = {
                    "byCategory": "bar",
                    "byStatus": "pie",
                    "monthlyTrend": "line",
                    "netLossAndRecovery": "bar",
                    "topFinancialImpacts": "bar",
                    "incidentsByEventType": "pie",
                    "incidentsByFinancialImpact": "pie",
                }
                chart_type = header_config.get("chartType") or header_config.get("chart_type")
                if chart_type not in {"bar", "line", "pie"}:
                    chart_type = default_type_by_card.get(card_type, "bar")
                header_config["chart_type"] = chart_type

                # Extract chart data from rows
                chart_labels: List[str] = []
                chart_values: List[float] = []
                for row in data_rows:
                    if len(row) >= 2:
                        chart_labels.append(str(row[0]))
                        try:
                            chart_values.append(float(row[1]))
                        except Exception:
                            chart_values.append(0)
                if chart_labels and chart_values:
                    header_config["chart_data"] = {"labels": chart_labels, "values": chart_values}
                return generate_excel_report(columns, data_rows, header_config)

            # TABLE EXPORT
            elif only_overall_table:
                table_rows = []
                if card_type == 'overallStatuses':
                    table_rows = incidents_data.get('overallStatuses') or incidents_data.get('statusOverview') or []
                else:
                    table_rows = incidents_data.get(card_type) or []

                if isinstance(table_rows, list) and len(table_rows) > 0:
                    first_item = table_rows[0]
                    if isinstance(first_item, dict):
                        raw_keys = list(first_item.keys())
                        columns = ['#'] + [format_column_name(k) for k in raw_keys]
                        for i, row in enumerate(table_rows, 1):
                            values = [str(row.get(k, '')) for k in raw_keys]
                            data_rows.append([str(i)] + values)
                    elif isinstance(first_item, (list, tuple)):
                        num_cols = len(first_item)
                        columns = ['#'] + [f'C{idx+1}' for idx in range(num_cols)]
                        for i, row in enumerate(table_rows, 1):
                            vals = [str(v) for v in (row if isinstance(row, (list, tuple)) else [row])]
                            data_rows.append([str(i)] + vals)
                    else:
                        columns = ['#', 'Value']
                        data_rows = [["1", str(first_item)]]
                else:
                    columns = ['#', 'Value']
                    data_rows = [["1", 'No data available']]
                return generate_excel_report(columns, data_rows, header_config)

            # CARD SUMMARY EXPORT
            elif only_card and card_type:
                if isinstance(data, list) and data:
                    first_item = data[0]
                    if isinstance(first_item, dict):
                        raw_keys = list(first_item.keys())
                        columns = ['#'] + [format_column_name(k) for k in raw_keys]
                        for i, item in enumerate(data, 1):
                            values = [str(item.get(k, 'N/A')) for k in raw_keys]
                            data_rows.append([str(i)] + values)
                    else:
                        columns = ['#', 'Value']
                        data_rows = [["1", str(first_item)]]
                elif isinstance(data, dict):
                    columns = ["Metric", "Value"]
                    data_rows = [[key, str(value)] for key, value in data.items()]
                else:
                    columns = ["Metric", "Value"]
                    data_rows = [["No data available", "N/A"]]
                return generate_excel_report(columns, data_rows, header_config)

            # DEFAULT
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = header_config.get('title', 'Incidents Report')
                ws['A1'] = 'Incidents Dashboard Report'
                ws['A2'] = 'Generated Successfully'
                from io import BytesIO
                output = BytesIO()
                wb.save(output)
                return output.getvalue()

        except Exception as e:
            print(f"Error generating incidents Excel: {e}")
            raise e

    async def generate_kris_excel(self, kris_data: Dict[str, Any], start_date: str, end_date: str, header_config: Dict[str, Any], card_type: str = None, only_card: bool = False, only_overall_table: bool = False, only_chart: bool = False) -> bytes:
        """Generate KRI Excel report mirroring incidents behavior."""
        try:
            from routes.route_utils import generate_excel_report, write_debug
            import re
            write_debug(f"Generating KRIs Excel report for {start_date} to {end_date}")
            write_debug(f"card_type={card_type}, only_card={only_card}, only_overall_table={only_overall_table}, only_chart={only_chart}")

            data = kris_data.get(card_type) if card_type else kris_data.get('list')
            data = data or []
            write_debug(f"KRIs Excel export - card_type={card_type}, data type={type(data)}, len={len(data) if isinstance(data, list) else 'N/A'}")

            columns: List[str] = []
            data_rows: List[List[str]] = []

            def format_column_name(key: str) -> str:
                return re.sub(r'[_]|([a-z])([A-Z])', r'\1 \2', str(key)).title()

            # CHART EXPORT
            if only_chart and card_type:
                # Special handling for stacked monthly assessment chart
                if card_type == "kriMonthlyAssessment" and isinstance(data, list) and data:
                    # Transform from long format to wide format (pivot)
                    grouped: dict = {}
                    assessment_levels = set()
                    
                    def parse_month(month_val):
                        """Parse month to datetime object and return formatted string"""
                        try:
                            from datetime import datetime
                            dt = None
                            if isinstance(month_val, str):
                                # Try parsing various date formats
                                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d', '%Y-%m-%dT%H:%M:%S.%f']:
                                    try:
                                        dt = datetime.strptime(month_val.split('.')[0], fmt)
                                        break
                                    except:
                                        continue
                            elif hasattr(month_val, 'strftime'):
                                # Already a datetime object
                                dt = month_val
                            
                            if dt:
                                return dt, dt.strftime('%b %Y')  # Format as "Jan 2025"
                            return None, str(month_val)
                        except:
                            return None, str(month_val)
                    
                    # Use dict with date key for proper sorting
                    month_dict = {}  # Maps formatted string to date object for sorting
                    
                    for item in data:
                        # Get month from createdAt, month, or month_year
                        month = item.get('month') or item.get('month_year') or item.get('createdAt')
                        if month:
                            dt, month_str = parse_month(month)
                            
                            assessment = item.get('assessment') or item.get('level') or 'Unknown'
                            count = float(item.get('count', 0))
                            
                            assessment_levels.add(assessment)
                            month_dict[month_str] = dt  # Store date for sorting
                            
                            if month_str not in grouped:
                                grouped[month_str] = {}
                            grouped[month_str][assessment] = (grouped[month_str].get(assessment, 0) + count)
                    
                    # Create columns: Month + each assessment level
                    sorted_levels = sorted(list(assessment_levels))
                    columns = ["Month"] + sorted_levels
                    
                    # Sort months by date value (not string)
                    from datetime import datetime
                    sorted_months = sorted(grouped.keys(), key=lambda m: month_dict.get(m) or datetime(1970, 1, 1))
                    chart_labels: List[str] = []
                    chart_series = []
                    
                    for month in sorted_months:
                        row = [month]
                        for level in sorted_levels:
                            value = grouped[month].get(level, 0)
                            row.append(str(value))
                        data_rows.append(row)
                        chart_labels.append(month)
                    
                    # For stacked charts, create series data
                    for level in sorted_levels:
                        series_values = []
                        for month in sorted_months:
                            value = grouped[month].get(level, 0)
                            series_values.append(float(value))
                        chart_series.append({"name": level, "values": series_values})
                    
                    header_config["chart_data"] = {
                        "labels": chart_labels,
                        "series": chart_series
                    }
                    header_config["chart_type"] = "bar"
                    header_config["stacked"] = True
                else:
                    # Default handling for other charts
                    if isinstance(data, list) and data:
                        first_item = data[0]
                        if isinstance(first_item, dict):
                            keys = list(first_item.keys())
                            if len(keys) >= 2:
                                key1 = keys[0]
                                key2 = keys[-1]
                                columns = [format_column_name(key1), format_column_name(key2)]
                                for item in data:
                                    name = item.get(key1, "N/A")
                                    value = item.get(key2, 0)
                                    data_rows.append([name, str(value)])
                            else:
                                key1 = keys[0]
                                columns = [format_column_name(key1), "Value"]
                                for item in data:
                                    name = item.get(key1, "N/A")
                                    data_rows.append([name, 0])
                        else:
                            data_rows = [[str(first_item), "0"]]
                            columns = ["Label", "Value"]
                    else:
                        data_rows = [["No data available", "0"]]
                        columns = ["Label", "Value"]

                    default_type_by_card = {
                        "krisByStatus": "pie",
                        "krisByLevel": "pie",
                        "breachedKRIsByDepartment": "bar",
                        "kriAssessmentCount": "bar",
                        "kriOverdueStatusCounts": "pie",
                        "kriMonthlyAssessment": "bar",
                    }
                    chart_type = header_config.get("chartType") or header_config.get("chart_type")
                    if chart_type not in {"bar", "line", "pie"}:
                        chart_type = default_type_by_card.get(card_type, "bar")
                    header_config["chart_type"] = chart_type

                    # Extract chart data from rows
                    chart_labels = []
                    chart_values = []
                    for row in data_rows:
                        if len(row) >= 2:
                            chart_labels.append(str(row[0]))
                            try:
                                chart_values.append(float(row[1]))
                            except Exception:
                                chart_values.append(0)
                    if chart_labels and chart_values:
                        header_config["chart_data"] = {"labels": chart_labels, "values": chart_values}
                
                return generate_excel_report(columns, data_rows, header_config)

            # TABLE EXPORT
            elif only_overall_table:
                table_rows = kris_data.get(card_type) or []

                if isinstance(table_rows, list) and len(table_rows) > 0:
                    first_item = table_rows[0]
                    if isinstance(first_item, dict):
                        raw_keys = list(first_item.keys())
                        columns = ['#'] + [format_column_name(k) for k in raw_keys]
                        for i, row in enumerate(table_rows, 1):
                            values = [str(row.get(k, '')) for k in raw_keys]
                            data_rows.append([str(i)] + values)
                    elif isinstance(first_item, (list, tuple)):
                        num_cols = len(first_item)
                        columns = ['#'] + [f'C{idx+1}' for idx in range(num_cols)]
                        for i, row in enumerate(table_rows, 1):
                            vals = [str(v) for v in (row if isinstance(row, (list, tuple)) else [row])]
                            data_rows.append([str(i)] + vals)
                    else:
                        columns = ['#', 'Value']
                        data_rows = [["1", str(first_item)]]
                else:
                    columns = ['#', 'Value']
                    data_rows = [["1", 'No data available']]
                
                write_debug(f"About to call generate_excel_report for KRIs table")
                result = generate_excel_report(columns, data_rows, header_config)
                write_debug(f"KRIs Excel report generated, returning {len(result) if result else 0} bytes")
                return result

            # CARD SUMMARY EXPORT
            elif only_card and card_type:
                write_debug(f"Generating KRIs Excel report for card {card_type}")
                # Handle both dict and list data
                if isinstance(data, list) and data:
                    first_item = data[0] if data else {}
                    if isinstance(first_item, dict):
                        raw_keys = list(first_item.keys())
                        columns = ['#'] + [format_column_name(k) for k in raw_keys]
                        for i, item in enumerate(data, 1):
                            values = [str(item.get(k, 'N/A')) for k in raw_keys]
                            data_rows.append([str(i)] + values)
                    else:
                        columns = ['#', 'Value']
                        data_rows = [[str(i+1), str(v)] for i, v in enumerate(data)]
                elif isinstance(data, dict):
                    columns = ["Metric", "Value"]
                    data_rows = [[key, str(value)] for key, value in data.items()]
                else:
                    columns = ["Metric", "Value"]
                    data_rows = [["No data available", "N/A"]]
                return generate_excel_report(columns, data_rows, header_config)

            # DEFAULT simple workbook if no specific mode
            else:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = header_config.get('title', 'KRIs Report')
                ws['A1'] = 'KRIs Dashboard Report'
                ws['A2'] = 'Generated Successfully'
                from io import BytesIO
                output = BytesIO()
                wb.save(output)
                return output.getvalue()
                
        except Exception as e:
            print(f"Error generating KRIs Excel: {e}")
            raise e
   
    async def generate_risks_excel(self, risks_data: Dict[str, Any], start_date: str, end_date: str, header_config: Dict[str, Any], card_type: str = None, only_card: bool = False, only_overall_table: bool = False, only_chart: bool = False) -> bytes:
        """Generate risks Excel report mirroring controls Excel behavior."""
        try:
            from routes.route_utils import generate_excel_report, write_debug
            write_debug(f"Generating risks Excel report for {start_date} to {end_date}")
            write_debug(f"card_type={card_type}, only_card={only_card}, only_overall_table={only_overall_table}, only_chart={only_chart}")
            write_debug(f"risks_data: {risks_data}")
            import re

            data = risks_data.get(card_type, []) if risks_data else []
            write_debug(f"Risks Excel export - card_type={card_type}, data type={type(data)}, len={len(data) if isinstance(data, list) else 'N/A'}, only_card={only_card}, only_overall_table={only_overall_table}, only_chart={only_chart}")

            columns: List[str] = []
            data_rows: List[List[str]] = []

            def format_column_name(key: str) -> str:
                write_debug(f"Formatting column name: {key}")
                return re.sub(r'[_]|([a-z])([A-Z])', r'\1 \2', str(key)).title()


            write_debug(f"Formatting column name: {format_column_name('code')}")

            # CHART EXPORT (same style as controls)
            if only_chart and card_type:
                if isinstance(data, list) and data:
                    first_item = data[0]
                    if isinstance(first_item, dict):
                        keys = list(first_item.keys())
                        # Special handling for charts with multiple value columns
                        if card_type == "createdDeletedRisksPerQuarter" and len(keys) >= 3:
                            # For created/deleted charts, include all numeric columns
                            label_key = "name"
                            value_keys = [k for k in keys if k != label_key]
                            write_debug(f"[RISKS EXCEL] Multi-column chart: label_key={label_key}, value_keys={value_keys}")
                            columns = [format_column_name(label_key)] + [format_column_name(k) for k in value_keys]
                            for item in data:
                                name = item.get(label_key, "N/A")
                                row = [name]
                                for vk in value_keys:
                                    val = item.get(vk, 0)
                                    row.append(str(val))
                                data_rows.append(row)
                            write_debug(f"[RISKS EXCEL] Generated {len(data_rows)} rows, columns={columns}")
                        elif len(keys) >= 2:
                            key1 = keys[0]
                            key2 = keys[-1]
                            columns = [format_column_name(key1), format_column_name(key2)]
                            for item in data:
                                name = item.get(key1, "N/A")
                                value = item.get(key2, 0)
                                data_rows.append([name, str(value)])
                        else:
                            key1 = keys[0]
                            columns = [format_column_name(key1), "Value"]
                            for item in data:
                                name = item.get(key1, "N/A")
                                data_rows.append([name, 0])
                    else:
                        data_rows = [[str(first_item), "0"]]
                        columns = ["Label", "Value"]
                else:
                    data_rows = [["No data available", "0"]]
                    columns = ["Label", "Value"]

                # Default chart types for risks
                default_type_by_card = {
                    "risksByCategory": "bar",
                    "risksByEventType": "pie",
                    "createdDeletedRisksPerQuarter": "bar",
                    "quarterlyRiskCreationTrends": "line",
                    "riskApprovalStatusDistribution": "pie",
                    "riskDistributionByFinancialImpact": "pie",
                }
                
                # Priority: renderType/chartType from query > cardType default > "bar"
                chart_type = header_config.get("chartType") or header_config.get("chart_type")
                if chart_type and chart_type in {"bar", "line", "pie"}:
                    write_debug(f"Using provided chart type: {chart_type}")
                else:
                    chart_type = default_type_by_card.get(card_type, "bar")
                    write_debug(f"Using default chart type for {card_type}: {chart_type}")
                
                header_config["chart_type"] = chart_type

                # Extract chart data from rows
                chart_labels: List[str] = []
                chart_values: List[float] = []
                for row in data_rows:
                    if len(row) >= 2:
                        chart_labels.append(str(row[0]))
                        try:
                            chart_values.append(float(row[1]))
                        except Exception:
                            chart_values.append(0)
                if chart_labels and chart_values:
                    header_config["chart_data"] = {"labels": chart_labels, "values": chart_values}
                return generate_excel_report(columns, data_rows, header_config)

            # TABLE EXPORT (same dynamic building as controls)
            elif only_overall_table and card_type:
                if isinstance(data, list) and len(data) > 0:
                    first_item = data[0]
                    if isinstance(first_item, dict):
                        # Special formatting for allRisks
                        if card_type == 'allRisks':
                            columns = ['#', 'Risk Name', 'Risk Description', 'Event', 'Inherent Value', 'Frequency', 'Financial Impact']
                            def map_frequency(val):
                                mapping = {1: 'Once in Three Years', 2: 'Annually', 3: 'Half Yearly', 4: 'Quarterly', 5: 'Monthly'}
                                try:
                                    num = int(val)
                                except Exception:
                                    return str(val)
                                return mapping.get(num, str(val))
                            def map_financial(val):
                                mapping = {1: '0 - 10,000', 2: '10,000 - 100,000', 3: '100,000 - 1,000,000', 4: '1,000,000 - 10,000,000', 5: '> 10,000,000'}
                                try:
                                    num = int(val)
                                except Exception:
                                    return str(val)
                                return mapping.get(num, str(val))
                            data_rows = []
                            for i, row in enumerate(data, 1):
                                data_rows.append([
                                    str(i),
                                    str(row.get('RiskName', '')),
                                    str(row.get('RiskDesc', '')),
                                    str(row.get('RiskEventName', 'Unknown')),
                                    str(row.get('InherentValue', '')),
                                    map_frequency(row.get('InherentFrequency', '')),
                                    map_financial(row.get('InherentFinancialValue', '')),
                                ])
                        else:
                            raw_keys = list(first_item.keys())
                            columns = ['#'] + [format_column_name(k) for k in raw_keys]
                            data_rows = []
                            for i, row in enumerate(data, 1):
                                if isinstance(row, dict):
                                    values = [str(row.get(k, '')) for k in raw_keys]
                                    data_rows.append([str(i)] + values)
                                elif isinstance(row, (list, tuple)):
                                    vals = [str(v) for v in row]
                                    data_rows.append([str(i)] + vals)
                                else:
                                    data_rows.append([str(i), str(row)])
                    elif isinstance(first_item, (list, tuple)):
                        num_cols = len(first_item)
                        columns = ['#'] + [f'C{idx+1}' for idx in range(num_cols)]
                        data_rows = []
                        for i, row in enumerate(data, 1):
                            vals = [str(v) for v in (row if isinstance(row, (list, tuple)) else [row])]
                            data_rows.append([str(i)] + vals)
                    else:
                        columns = ['#', 'Value']
                        data_rows = [[str(i+1), str(v)] for i, v in enumerate(data)]
                else:
                    columns = ['#', 'Value']
                    data_rows = [['1', 'No data available']]
                return generate_excel_report(columns, data_rows, header_config)

            # CARD SUMMARY EXPORT
            elif only_card and card_type:
                write_debug(f"Generating risks Excel report for mmmmm {start_date} to {end_date}")
                write_debug(f"card_type={card_type}, only_card={only_card}, only_overall_table={only_overall_table}, only_chart={only_chart}")
                write_debug(f"data: {data}")
               
                if isinstance(data, list) and data:
                    first_item = data[0]
                    if isinstance(first_item, dict):
                        raw_keys = list(first_item.keys())
                        columns = ['#'] + [format_column_name(k) for k in raw_keys]
                        data_rows = []
                        for i, item in enumerate(data, 1):
                            if isinstance(item, dict):
                                values = [str(item.get(k, 'N/A')) for k in raw_keys]
                                data_rows.append([str(i)] + values)
                            elif isinstance(item, (list, tuple)):
                                vals = [str(v) for v in item]
                                data_rows.append([str(i)] + vals)
                            else:
                                data_rows.append([str(i), str(item)])
                    elif isinstance(first_item, str):
                        columns = ["#", "Risk Name"]
                        for i, item in enumerate(data, 1):
                            data_rows.append([str(i), str(item)])
                    else:
                        columns = ['#', 'Value']
                        data_rows = [["1", str(first_item)]]
                elif isinstance(data, dict):
                    columns = ["Metric", "Value"]
                    data_rows = [[key, str(value)] for key, value in data.items()]
                else:
                    columns = ["Metric", "Value"]
                    data_rows = [["No data available", "N/A"]]
                return generate_excel_report(columns, data_rows, header_config)

            # DEFAULT simple workbook if no specific mode
            else:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = header_config.get('title', 'Risks Report')
                ws['A1'] = 'Risks Dashboard Report'
                ws['A2'] = 'Generated Successfully'
                from io import BytesIO
                output = BytesIO()
                wb.save(output)
                return output.getvalue()
                
        except Exception as e:
            print(f"Error generating risks Excel: {e}")
            raise e