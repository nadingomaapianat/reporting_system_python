from fastapi import FastAPI, HTTPException, Depends, BackgroundTasks, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from shared.dashboard_controller import router as dashboard_router
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.graph_objects as go
import plotly.express as px
from plotly.utils import PlotlyJSONEncoder
import plotly.offline as pyo
import json
import io
import base64
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Union
import os
from pathlib import Path
import asyncio
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, BaseDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, KeepInFrame, Frame, PageTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics import renderPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
try:
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
except Exception:
    XLImage = None
    PILImage = None
import sqlite3
from sqlalchemy import create_engine, text
import redis
import pickle

# Optional Arabic support
try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    ARABIC_AVAILABLE = True
except Exception:
    ARABIC_AVAILABLE = False

# Try to register an Arabic-capable font if available in ./fonts or Windows fonts
ARABIC_FONT_NAME = None
if ARABIC_AVAILABLE:
    try:
        candidates = []
        # Project-local fonts directory
        fonts_dir = Path(__file__).parent / 'fonts'
        candidates += [fonts_dir / x for x in [
            'NotoNaskhArabic-Regular.ttf', 'Amiri-Regular.ttf', 'NotoSansArabic-Regular.ttf',
            'DejaVuSans.ttf', 'Scheherazade-Regular.ttf'
        ]]
        # Common Windows fonts that support Arabic
        win_fonts = Path('C:/Windows/Fonts')
        if win_fonts.exists():
            candidates += [win_fonts / x for x in [
                'NotoNaskhArabic-Regular.ttf', 'Amiri-Regular.ttf', 'arialuni.ttf',
                'Tahoma.ttf', 'ARIAL.TTF', 'Times New Roman.ttf', 'segoeui.ttf',
                'trado.ttf',  # Traditional Arabic
                'simpbdo.ttf' # Simplified Arabic Bold (fallback)
            ]]
        fpath = next((p for p in candidates if p and p.exists()), None)
        if fpath:
            pdfmetrics.registerFont(TTFont('ArabicFont', str(fpath)))
            ARABIC_FONT_NAME = 'ArabicFont'
    except Exception:
        ARABIC_FONT_NAME = None

def shape_text_for_arabic(text: str) -> str:
    if not text:
        return text
    if ARABIC_AVAILABLE:
        try:
            reshaped = arabic_reshaper.reshape(text)
            return get_display(reshaped)
        except Exception:
            return text
    return text

app = FastAPI(title="Unified Dashboard & Reporting API", version="2.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "https://reporting-system-backend.pianat.ai",
        "http://127.0.0.1:3002",
        "http://localhost:3002"
    ],
    allow_origin_regex=r"^http://(localhost|127\\.0\\.0\\.1):(3000|3002)$",
    allow_credentials=True,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"]
)

# Include dashboard router
app.include_router(dashboard_router, prefix="/api", tags=["dashboard"])

# Create reports directory
reports_dir = Path("reports")
reports_dir.mkdir(exist_ok=True)

# Banking report configuration
BANK_CONFIG = {
    "name": "PIANAT.AI",
    "logo": "🏦",  # You can replace with actual logo path
    "address": "King Abdulaziz Road, Riyadh, Saudi Arabia",
    "phone": "+966 11 402 9000",
    "website": "www.alrajhibank.com.sa",
    "report_footer": "Confidential Banking Report - Internal Use Only"
}

# Redis connection for caching
try:
    redis_client = redis.Redis(host='localhost', port=6379, db=0, decode_responses=True)
    redis_client.ping()
    REDIS_AVAILABLE = True
except:
    REDIS_AVAILABLE = False
    print("Redis not available, using in-memory cache")

# Database setup
DATABASE_URL = "sqlite:///./dashboard.db"
engine = create_engine(DATABASE_URL)

# Cache for storing computed results
cache = {}

# Sample data generation
def generate_sample_data():
    """Generate sample data for dashboard"""
    np.random.seed(42)
    dates = pd.date_range(start='2024-01-01', end='2024-12-31', freq='D')
    
    data = {
        'date': dates,
        'sales': np.random.normal(1000, 200, len(dates)).cumsum(),
        'users': np.random.poisson(50, len(dates)).cumsum(),
        'revenue': np.random.normal(5000, 1000, len(dates)).cumsum(),
        'conversion_rate': np.random.beta(2, 8, len(dates)) * 100,
        'category': np.random.choice(['Electronics', 'Clothing', 'Books', 'Home'], len(dates)),
        'region': np.random.choice(['North', 'South', 'East', 'West'], len(dates))
    }
    
    return pd.DataFrame(data)

# Global data store
df = generate_sample_data()

@app.get("/")
async def root():
    return {"message": "Dashboard Analytics API", "status": "running"}

@app.get("/api/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

@app.get("/api/dashboard/overview")
async def get_dashboard_overview():
    """Get dashboard overview statistics"""
    latest_data = df.iloc[-1]
    previous_data = df.iloc[-2] if len(df) > 1 else latest_data
    
    return {
        "total_sales": int(latest_data['sales']),
        "total_users": int(latest_data['users']),
        "total_revenue": int(latest_data['revenue']),
        "conversion_rate": round(latest_data['conversion_rate'], 2),
        "sales_growth": round(((latest_data['sales'] - previous_data['sales']) / previous_data['sales']) * 100, 2),
        "user_growth": round(((latest_data['users'] - previous_data['users']) / previous_data['users']) * 100, 2),
        "revenue_growth": round(((latest_data['revenue'] - previous_data['revenue']) / previous_data['revenue']) * 100, 2),
        "CompanyEstablishedAt": None
    }

# GRC Dashboard Endpoints
@app.get("/api/grc/controls")
async def get_controls_dashboard():
    """Get controls dashboard data"""
    # Mock data - replace with actual database queries
    return {
        "totalControls": 1250,
        "unmappedControls": 45,
        "pendingPreparer": 23,
        "pendingChecker": 18,
        "pendingReviewer": 12,
        "pendingAcceptance": 8,
        "controlsByDepartment": [
            {"department_name": "IT", "controls_count": 320},
            {"department_name": "Finance", "controls_count": 280},
            {"department_name": "Operations", "controls_count": 250},
            {"department_name": "Risk Management", "controls_count": 200},
            {"department_name": "Compliance", "controls_count": 200}
        ],
        "controlsByRiskResponse": [
            {"risk_response": "Accept", "count": 400},
            {"risk_response": "Mitigate", "count": 500},
            {"risk_response": "Transfer", "count": 200},
            {"risk_response": "Avoid", "count": 150}
        ],
        "overallStatuses": [
            {
                "control_name": "Access Control Management",
                "acceptance_status": "approved",
                "checker_status": "approved",
                "preparer_status": "approved",
                "reviewer_status": "approved"
            },
            {
                "control_name": "Data Backup Procedures",
                "acceptance_status": "pending",
                "checker_status": "approved",
                "preparer_status": "approved",
                "reviewer_status": "approved"
            }
        ]
    }

@app.get("/api/grc/incidents")
async def get_incidents_dashboard():
    """Get incidents dashboard data"""
    # Mock data - replace with actual database queries
    return {
        "totalIncidents": 89,
        "pendingPreparer": 15,
        "pendingChecker": 12,
        "pendingReviewer": 8,
        "pendingAcceptance": 5,
        "incidentsByCategory": [
            {"category_name": "Cybersecurity", "count": 25},
            {"category_name": "Operational", "count": 20},
            {"category_name": "Financial", "count": 18},
            {"category_name": "Compliance", "count": 15},
            {"category_name": "Technology", "count": 11}
        ],
        "incidentsByStatus": [
            {"status": "Investigating", "count": 30},
            {"status": "Resolved", "count": 45},
            {"status": "Pending Review", "count": 14}
        ],
        "topFinancialImpacts": [
            {
                "incident_id": 1,
                "financial_impact_name": "Data Breach",
                "function_name": "IT",
                "net_loss": 500000
            },
            {
                "incident_id": 2,
                "financial_impact_name": "System Downtime",
                "function_name": "Operations",
                "net_loss": 250000
            }
        ],
        "totalRecoveryAmounts": [
            {
                "incident_id": 1,
                "financial_impact_name": "Data Breach",
                "function_name": "IT",
                "recovery_amount": 100000
            }
        ],
        "netLossAndRecovery": [
            {
                "incident_title": "Major Data Breach",
                "net_loss": 500000,
                "recovery_amount": 100000,
                "function_name": "IT"
            },
            {
                "incident_title": "System Outage",
                "net_loss": 250000,
                "recovery_amount": 50000,
                "function_name": "Operations"
            }
        ],
        "monthlyTrend": [
            {"month_year": "Jan 2024", "total_loss": 750000, "incident_count": 12},
            {"month_year": "Feb 2024", "total_loss": 600000, "incident_count": 10},
            {"month_year": "Mar 2024", "total_loss": 450000, "incident_count": 8}
        ]
    }

@app.get("/api/grc/kris")
async def get_kris_dashboard():
    """Get KRIs dashboard data"""
    # Mock data - replace with actual database queries
    return {
        "totalKRIs": 45,
        "breachedKRIs": 8,
        "pendingPreparer": 5,
        "pendingChecker": 3,
        "pendingReviewer": 2,
        "pendingAcceptance": 1,
        "lowKRIs": 15,
        "mediumKRIs": 20,
        "highKRIs": 10,
        "kriHealth": [
            {
                "kri_code": "KRI001",
                "kri_name": "System Uptime",
                "function": "IT",
                "status": "High",
                "frequency": "Daily",
                "threshold": 99.5,
                "workflow_status": "Complete"
            },
            {
                "kri_code": "KRI002",
                "kri_name": "Transaction Volume",
                "function": "Operations",
                "status": "Medium",
                "frequency": "Daily",
                "threshold": 10000,
                "workflow_status": "Pending Checker"
            }
        ],
        "breachedKRIsByDepartment": [
            {"kri_id": 1, "kri_name": "System Uptime", "function_name": "IT", "breached_count": 3},
            {"kri_id": 2, "kri_name": "Transaction Volume", "function_name": "Operations", "breached_count": 2}
        ],
        "kriAssessmentCount": [
            {"kri_id": 1, "function_name": "IT", "assessment": "High", "kri_name": "System Uptime"},
            {"kri_id": 2, "function_name": "Operations", "assessment": "Medium", "kri_name": "Transaction Volume"}
        ],
        "kriStatus": [
            {
                "id": 1,
                "code": "KRI001",
                "kri_name": "System Uptime",
                "threshold": 99.5,
                "is_ascending": True,
                "kri_level": "High",
                "status": "Active",
                "created_at": "2024-01-01",
                "updated_at": "2024-01-15",
                "function_name": "IT"
            }
        ]
    }

@app.get("/api/grc/risks")
async def get_risks_dashboard():
    """Get risks dashboard data"""
    # Mock data - replace with actual database queries
    return {
        "totalRisks": 156,
        "risksByCategory": [
            {"category": "Operational", "risk_id": 1},
            {"category": "Financial", "risk_id": 2},
            {"category": "Cybersecurity", "risk_id": 3},
            {"category": "Compliance", "risk_id": 4}
        ],
        "risksByEventType": [
            {"event_type": "System Failure", "risk_id": 1},
            {"event_type": "Data Breach", "risk_id": 2},
            {"event_type": "Regulatory Violation", "risk_id": 3}
        ],
        "inherentVsResidual": [
            {
                "risk_id": 1,
                "risk_name": "System Downtime Risk",
                "inherent_value": 8,
                "residual_value": 4,
                "created_at": "2024-01-01"
            },
            {
                "risk_id": 2,
                "risk_name": "Data Security Risk",
                "inherent_value": 9,
                "residual_value": 3,
                "created_at": "2024-01-02"
            }
        ],
        "riskLevels": [
            {"level": "High", "count": 25},
            {"level": "Medium", "count": 80},
            {"level": "Low", "count": 51}
        ],
        "riskTrends": [
            {"month": "Jan 2024", "total_risks": 150, "new_risks": 15, "mitigated_risks": 5},
            {"month": "Feb 2024", "total_risks": 155, "new_risks": 12, "mitigated_risks": 7},
            {"month": "Mar 2024", "total_risks": 156, "new_risks": 8, "mitigated_risks": 7}
        ]
    }

# GRC Export Endpoints
@app.post("/api/grc/controls/export")
async def export_controls_report(format: str = "pdf", request_data: dict = None):
    """Export controls report in PDF or Excel format"""
    try:
        if format == "pdf":
            # Generate PDF report
            filename = f"controls_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = reports_dir / filename
            
            # Create PDF report (simplified version)
            doc = SimpleDocTemplate(str(filepath), pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title = Paragraph("Controls Dashboard Report", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Summary data
            data = [['Metric', 'Value'],
                   ['Total Controls', '1,250'],
                   ['Unmapped Controls', '45'],
                   ['Pending Preparer', '23'],
                   ['Pending Checker', '18'],
                   ['Pending Reviewer', '12'],
                   ['Pending Acceptance', '8']]
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            doc.build(story)
            
            return FileResponse(filepath, media_type='application/pdf', filename=filename)
            
        elif format == "excel":
            # Generate Excel report
            filename = f"controls_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = reports_dir / filename
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Controls Report"
            
            # Add headers
            headers = ['Metric', 'Value']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            
            # Add data
            data = [
                ['Total Controls', 1250],
                ['Unmapped Controls', 45],
                ['Pending Preparer', 23],
                ['Pending Checker', 18],
                ['Pending Reviewer', 12],
                ['Pending Acceptance', 8]
            ]
            
            for row, row_data in enumerate(data, 2):
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            wb.save(filepath)
            return FileResponse(filepath, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=filename)
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@app.post("/api/grc/incidents/export")
async def export_incidents_report(format: str = "pdf", request_data: dict = None):
    """Export incidents report in PDF or Excel format"""
    # Similar implementation for incidents
    return {"message": "Incidents export endpoint - implementation similar to controls"}

@app.post("/api/grc/kris/export")
async def export_kris_report(format: str = "pdf", request_data: dict = None):
    """Export KRIs report in PDF or Excel format"""
    # Similar implementation for KRIs
    return {"message": "KRIs export endpoint - implementation similar to controls"}

@app.post("/api/grc/risks/export")
async def export_risks_report(format: str = "pdf", request_data: dict = None):
    """Export risks report in PDF or Excel format"""
    # Similar implementation for risks
    return {"message": "Risks export endpoint - implementation similar to controls"}

@app.get("/api/charts/sales-trend")
async def get_sales_trend(days: int = 30):
    """Get sales trend data for charts"""
    recent_data = df.tail(days)
    
    return {
        "labels": recent_data['date'].dt.strftime('%Y-%m-%d').tolist(),
        "datasets": [
            {
                "label": "Sales",
                "data": recent_data['sales'].tolist(),
                "borderColor": "rgb(59, 130, 246)",
                "backgroundColor": "rgba(59, 130, 246, 0.1)"
            }
        ]
    }

@app.get("/api/charts/revenue-by-category")
async def get_revenue_by_category():
    """Get revenue breakdown by category"""
    category_data = df.groupby('category')['revenue'].sum().reset_index()
    
    return {
        "labels": category_data['category'].tolist(),
        "datasets": [
            {
                "label": "Revenue",
                "data": category_data['revenue'].tolist(),
                "backgroundColor": [
                    "rgba(255, 99, 132, 0.8)",
                    "rgba(54, 162, 235, 0.8)",
                    "rgba(255, 205, 86, 0.8)",
                    "rgba(75, 192, 192, 0.8)"
                ]
            }
        ]
    }

@app.get("/api/charts/conversion-funnel")
async def get_conversion_funnel():
    """Get conversion funnel data"""
    # Simulate funnel data
    funnel_data = {
        "labels": ["Visitors", "Leads", "Opportunities", "Customers"],
        "values": [10000, 2500, 500, 100],
        "colors": ["#FF6B6B", "#4ECDC4", "#45B7D1", "#96CEB4"]
    }
    
    return funnel_data

@app.get("/api/charts/geographic-distribution")
async def get_geographic_distribution():
    """Get geographic distribution data"""
    region_data = df.groupby('region').agg({
        'sales': 'sum',
        'users': 'sum',
        'revenue': 'sum'
    }).reset_index()
    
    return region_data.to_dict('records')

@app.get("/api/reports/generate")
async def generate_report(report_type: str = "summary", format: str = "json"):
    """Generate various types of reports"""
    
    if report_type == "summary":
        summary = {
            "total_records": len(df),
            "date_range": {
                "start": df['date'].min().strftime('%Y-%m-%d'),
                "end": df['date'].max().strftime('%Y-%m-%d')
            },
            "total_sales": int(df['sales'].sum()),
            "total_revenue": int(df['revenue'].sum()),
            "total_users": int(df['users'].sum()),
            "avg_conversion_rate": round(df['conversion_rate'].mean(), 2),
            "top_category": df.groupby('category')['revenue'].sum().idxmax(),
            "top_region": df.groupby('region')['sales'].sum().idxmax()
        }
        
        if format == "json":
            return summary
        elif format == "csv":
            # Generate CSV
            csv_data = df.to_csv(index=False)
            return {"csv_data": csv_data}
    
    elif report_type == "detailed":
        detailed_report = df.describe().to_dict()
        return detailed_report
    
    return {"error": "Invalid report type"}

@app.get("/api/analytics/trend-analysis")
async def get_trend_analysis(metric: str = "sales", period: str = "30d"):
    """Perform trend analysis on specified metric"""
    
    if metric not in ['sales', 'users', 'revenue', 'conversion_rate']:
        raise HTTPException(status_code=400, detail="Invalid metric")
    
    # Get data for the specified period
    if period == "7d":
        data = df.tail(7)
    elif period == "30d":
        data = df.tail(30)
    elif period == "90d":
        data = df.tail(90)
    else:
        data = df
    
    # Calculate trend
    values = data[metric].values
    x = np.arange(len(values))
    slope, intercept = np.polyfit(x, values, 1)
    trend_direction = "increasing" if slope > 0 else "decreasing"
    
    # Calculate percentage change
    if len(values) > 1:
        pct_change = ((values[-1] - values[0]) / values[0]) * 100
    else:
        pct_change = 0
    
    return {
        "metric": metric,
        "period": period,
        "trend_direction": trend_direction,
        "slope": float(slope),
        "percentage_change": round(pct_change, 2),
        "current_value": float(values[-1]),
        "data_points": len(values)
    }

@app.get("/api/analytics/correlation-matrix")
async def get_correlation_matrix():
    """Get correlation matrix between numeric metrics"""
    numeric_cols = ['sales', 'users', 'revenue', 'conversion_rate']
    correlation_matrix = df[numeric_cols].corr()
    
    return {
        "matrix": correlation_matrix.to_dict(),
        "labels": numeric_cols
    }

# ==================== ADVANCED REPORTING SYSTEM ====================

class ReportGenerator:
    """Advanced report generation with multiple formats"""
    
    @staticmethod
    def generate_pdf_report(data: dict, report_type: str = "summary") -> bytes:
        """Generate PDF report using ReportLab"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph(f"{report_type.title()} Report", title_style))
        story.append(Spacer(1, 20))
        
        # Add data as table
        if report_type == "summary":
            table_data = [["Metric", "Value", "Growth"]]
            for key, value in data.items():
                if isinstance(value, dict) and 'value' in value and 'growth' in value:
                    table_data.append([key.replace('_', ' ').title(), str(value['value']), f"{value['growth']}%"])
                else:
                    table_data.append([key.replace('_', ' ').title(), str(value), "-"])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def generate_excel_report(data: dict, df_data: pd.DataFrame, report_type: str = "summary") -> bytes:
        """Generate Excel report with charts using openpyxl"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{report_type.title()} Report"
        
        # Add summary data
        if report_type == "summary":
            ws['A1'] = f"{report_type.title()} Report"
            ws['A1'].font = Font(size=16, bold=True)
            
            row = 3
            for key, value in data.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                if isinstance(value, dict) and 'value' in value:
                    ws[f'B{row}'] = value['value']
                else:
                    ws[f'B{row}'] = value
                row += 1
            
            # Add chart
            chart = BarChart()
            chart.title = "Key Metrics"
            chart.x_axis.title = "Metrics"
            chart.y_axis.title = "Values"
            
            data_ref = Reference(ws, min_col=2, min_row=3, max_row=row-1)
            categories_ref = Reference(ws, min_col=1, min_row=3, max_row=row-1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(categories_ref)
            
            ws.add_chart(chart, "E3")
        
        # Add detailed data sheet
        if not df_data.empty:
            ws2 = wb.create_sheet("Detailed Data")
            for r in dataframe_to_rows(df_data, index=False, header=True):
                ws2.append(r)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

# ==================== ENHANCED ANALYTICS ENDPOINTS ====================

@app.get("/api/analytics/advanced-metrics")
async def get_advanced_metrics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    group_by: str = "day"
):
    """Get advanced analytics metrics with filtering"""
    cache_key = f"advanced_metrics_{start_date}_{end_date}_{group_by}"
    
    # Check cache first
    if REDIS_AVAILABLE:
        cached = redis_client.get(cache_key)
        if cached:
            return json.loads(cached)
    elif cache_key in cache:
        return cache[cache_key]
    
    # Filter data
    filtered_df = df.copy()
    if start_date:
        filtered_df = filtered_df[filtered_df['date'] >= start_date]
    if end_date:
        filtered_df = filtered_df[filtered_df['date'] <= end_date]
    
    # Group by specified period
    if group_by == "day":
        grouped = filtered_df.groupby(filtered_df['date'].dt.date)
    elif group_by == "week":
        grouped = filtered_df.groupby(filtered_df['date'].dt.isocalendar().week)
    elif group_by == "month":
        grouped = filtered_df.groupby(filtered_df['date'].dt.to_period('M'))
    else:
        grouped = filtered_df.groupby(filtered_df['date'].dt.date)
    
    # Calculate advanced metrics
    metrics = grouped.agg({
        'sales': ['sum', 'mean', 'std'],
        'users': ['sum', 'mean', 'std'],
        'revenue': ['sum', 'mean', 'std'],
        'conversion_rate': ['mean', 'std']
    }).round(2)
    
    # Calculate growth rates
    sales_growth = ((metrics[('sales', 'sum')].pct_change() * 100).fillna(0)).round(2)
    revenue_growth = ((metrics[('revenue', 'sum')].pct_change() * 100).fillna(0)).round(2)
    
    result = {
        "period": group_by,
        "date_range": {
            "start": str(filtered_df['date'].min().date()),
            "end": str(filtered_df['date'].max().date())
        },
        "metrics": metrics.to_dict(),
        "growth_rates": {
            "sales_growth": sales_growth.to_dict(),
            "revenue_growth": revenue_growth.to_dict()
        },
        "summary": {
            "total_sales": int(filtered_df['sales'].sum()),
            "total_revenue": int(filtered_df['revenue'].sum()),
            "total_users": int(filtered_df['users'].sum()),
            "avg_conversion_rate": round(filtered_df['conversion_rate'].mean(), 2),
            "data_points": len(filtered_df)
        }
    }
    
    # Cache result
    if REDIS_AVAILABLE:
        redis_client.setex(cache_key, 300, json.dumps(result, default=str))  # 5 min cache
    else:
        cache[cache_key] = result
    
    return result

@app.get("/api/analytics/predictive-analysis")
async def get_predictive_analysis(metric: str = "sales", days_ahead: int = 30):
    """Perform simple predictive analysis using linear regression"""
    if metric not in ['sales', 'users', 'revenue', 'conversion_rate']:
        raise HTTPException(status_code=400, detail="Invalid metric")
    
    # Get recent data
    recent_data = df.tail(90)  # Use last 90 days for prediction
    
    # Prepare data for linear regression
    X = np.arange(len(recent_data)).reshape(-1, 1)
    y = recent_data[metric].values
    
    # Fit linear regression
    from sklearn.linear_model import LinearRegression
    model = LinearRegression()
    model.fit(X, y)
    
    # Predict future values
    future_X = np.arange(len(recent_data), len(recent_data) + days_ahead).reshape(-1, 1)
    predictions = model.predict(future_X)
    
    # Calculate confidence interval (simplified)
    residuals = y - model.predict(X)
    std_error = np.std(residuals)
    confidence_interval = 1.96 * std_error  # 95% confidence
    
    # Generate future dates
    last_date = recent_data['date'].iloc[-1]
    future_dates = pd.date_range(start=last_date + timedelta(days=1), periods=days_ahead, freq='D')
    
    return {
        "metric": metric,
        "prediction_period": days_ahead,
        "model_accuracy": round(model.score(X, y), 4),
        "predictions": [
            {
                "date": date.strftime('%Y-%m-%d'),
                "predicted_value": round(pred, 2),
                "confidence_lower": round(pred - confidence_interval, 2),
                "confidence_upper": round(pred + confidence_interval, 2)
            }
            for date, pred in zip(future_dates, predictions)
        ],
        "trend": "increasing" if model.coef_[0] > 0 else "decreasing",
        "slope": round(model.coef_[0], 4)
    }

# ==================== REPORT GENERATION ENDPOINTS ====================

# Professional Banking Report Functions
def create_bank_header():
    """Create professional bank header for reports"""
    styles = getSampleStyleSheet()
    
    # Bank header style
    bank_style = ParagraphStyle(
        'BankHeader',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.darkblue,
        alignment=1,  # Center
        spaceAfter=20
    )
    
    # Bank info style
    info_style = ParagraphStyle(
        'BankInfo',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=1,  # Center
        spaceAfter=10
    )
    
    header_elements = [
        Paragraph(f"{BANK_CONFIG['logo']} {BANK_CONFIG['name']}", bank_style),
        Paragraph(BANK_CONFIG['address'], info_style),
        Paragraph(f"Tel: {BANK_CONFIG['phone']} | Web: {BANK_CONFIG['website']}", info_style),
        Spacer(1, 20)
    ]
    
    return header_elements

def create_bank_footer():
    """Create professional bank footer for reports"""
    styles = getSampleStyleSheet()
    
    footer_style = ParagraphStyle(
        'BankFooter',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1,  # Center
        spaceBefore=20
    )
    
    return [
        Spacer(1, 20),
        Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", footer_style),
        Paragraph(BANK_CONFIG['report_footer'], footer_style)
    ]

def generate_controls_pdf_report(controls_data: dict, start_date: str = None, end_date: str = None, header_config: dict = None):
    """Generate professional PDF report for Controls Dashboard"""
    buffer = io.BytesIO()
    # Use BaseDocTemplate so we can draw watermark on top via onPageEnd
    doc = BaseDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    
    # Get styles
    styles = getSampleStyleSheet()
    # Default Arabic-capable font for body text if available
    if ARABIC_FONT_NAME:
        try:
            styles['Normal'].fontName = ARABIC_FONT_NAME
        except Exception:
            pass
    
    # Create content
    content = []
    # Resolve table colors early so they are available to all sections
    try:
        table_header_bg = colors.HexColor(header_config.get('tableHeaderBgColor', '#1F4E79'))
    except Exception:
        table_header_bg = colors.darkblue
    try:
        table_body_bg = colors.HexColor(header_config.get('tableBodyBgColor', '#FFFFFF'))
    except Exception:
        table_body_bg = colors.white
    
    # Add configurable header if enabled
    if header_config and header_config.get('includeHeader', True):
        # Resolve style values from config
        try:
            font_color = colors.HexColor(header_config.get('fontColor', '#1F4E79'))
        except Exception:
            font_color = colors.darkblue
        try:
            bg_color = colors.HexColor(header_config.get('backgroundColor', '#FFFFFF'))
        except Exception:
            bg_color = colors.white
        try:
            border_color = colors.HexColor(header_config.get('borderColor', '#E5E7EB'))
        except Exception:
            border_color = colors.lightgrey
        border_width = int(header_config.get('borderWidth', 1) or 0)
        padding = int(header_config.get('padding', 20) or 0)
        margin = int(header_config.get('margin', 10) or 0)
        font_size_map = {'small': 14, 'medium': 18, 'large': 22}
        title_font_size = font_size_map.get(str(header_config.get('fontSize', 'medium')).lower(), 18)

        # Build header flowables
        header_flow = []

        # Optional logo placement (cap size and enforce scaling)
        logo_base64 = header_config.get('logoBase64')
        logo_position = str(header_config.get('logoPosition', 'left')).lower()
        if header_config.get('showLogo', True) and logo_base64:
            try:
                img_bytes = base64.b64decode(logo_base64.split(',')[-1])
                img_buf = io.BytesIO(img_bytes)
                logo_img = Image(img_buf)
                # Constrain logo size strictly using draw* properties
                desired_height = int(header_config.get('logoHeight', 36) or 36)
                if desired_height > 64:
                    desired_height = 64
                img_w = getattr(logo_img, 'imageWidth', None)
                img_h = getattr(logo_img, 'imageHeight', None)
                if img_w and img_h and img_h > 0:
                    scale = desired_height / float(img_h)
                    width = float(img_w) * scale
                    max_width_pts = 180.0  # ~2.5 inches
                    if width > max_width_pts:
                        scale = max_width_pts / float(img_w)
                        width = max_width_pts
                        desired_height = float(img_h) * scale
                    logo_img.drawWidth = width
                    logo_img.drawHeight = desired_height
                else:
                    # Fallback fixed size
                    logo_img.drawHeight = desired_height
                    logo_img.drawWidth = desired_height
                # Alignment
                if logo_position == 'center':
                    logo_img.hAlign = 'CENTER'
                elif logo_position == 'right':
                    logo_img.hAlign = 'RIGHT'
                else:
                    logo_img.hAlign = 'LEFT'
                header_flow.append(logo_img)
                header_flow.append(Spacer(1, 6))
            except Exception:
                # Ignore logo errors silently, continue with title
                pass

        # Optional bank information block directly under logo (inside header block)
        if header_config.get('showBankInfo', True):
            info_style = ParagraphStyle('BankInfo', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=1)
            if ARABIC_FONT_NAME:
                info_style.fontName = ARABIC_FONT_NAME
            # Align info same as logo position
            if logo_position == 'left':
                info_style.alignment = 0
            elif logo_position == 'right':
                info_style.alignment = 2
            else:
                info_style.alignment = 1
            bank_name = shape_text_for_arabic(str(header_config.get('bankName', BANK_CONFIG['name'])))
            header_flow.append(Paragraph(bank_name, info_style))
            bank_line_1 = shape_text_for_arabic(str(header_config.get('bankAddress', '')))
            bank_line_2 = shape_text_for_arabic(f"Tel: {header_config.get('bankPhone','')} | Web: {header_config.get('bankWebsite','')}")
            if bank_line_1:
                header_flow.append(Paragraph(bank_line_1, info_style))
            if bank_line_2:
                header_flow.append(Paragraph(bank_line_2, info_style))
            header_flow.append(Spacer(1, 6))

        # Title and optional subtitle
        # Shape Arabic text if needed
        raw_title = header_config.get('title', 'Controls Dashboard Report')
        raw_subtitle = header_config.get('subtitle', '')
        title_text = shape_text_for_arabic(raw_title)
        subtitle = shape_text_for_arabic(raw_subtitle)

        if subtitle:
            title_text += f"<br/><font size='12' color='gray'>{subtitle}</font>"

        if start_date and end_date:
            title_text += f"<br/><font size='10' color='gray'>({start_date} to {end_date})</font>"
        elif start_date:
            title_text += f"<br/><font size='10' color='gray'>(From {start_date})</font>"

        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=title_font_size,
            textColor=font_color,
            alignment=1,
            spaceAfter=12
        )
        if ARABIC_FONT_NAME:
            title_style.fontName = ARABIC_FONT_NAME
        header_flow.append(Paragraph(title_text, title_style))

        # Wrap header flowables in a container table to apply background/border/padding
        # Keep header within a bounded frame to avoid overflow
        kif = KeepInFrame(doc.width, 1.5*inch, header_flow, mode='shrink')
        header_container = Table([[kif]], colWidths=[None])
        tbl_style = [
            ('BACKGROUND', (0, 0), (-1, -1), bg_color),
            ('LEFTPADDING', (0, 0), (-1, -1), padding),
            ('RIGHTPADDING', (0, 0), (-1, -1), padding),
            ('TOPPADDING', (0, 0), (-1, -1), padding),
            ('BOTTOMPADDING', (0, 0), (-1, -1), padding),
        ]
        if border_width > 0:
            tbl_style.extend([
                ('BOX', (0, 0), (-1, -1), border_width, border_color),
            ])
        header_container.setStyle(TableStyle(tbl_style))

        # Apply external margin via Spacer
        if margin > 0:
            content.append(Spacer(1, margin))
        content.append(header_container)
        if margin > 0:
            content.append(Spacer(1, margin))

        # (Bank info already added under logo inside header block)
    
    # Executive Summary (skip if scoped export)
    if not header_config.get('scoped', False):
        exec_heading = ParagraphStyle('ExecHeading', parent=styles['Heading2'])
        if ARABIC_FONT_NAME:
            exec_heading.fontName = ARABIC_FONT_NAME
        content.append(Paragraph(shape_text_for_arabic("Executive Summary"), exec_heading))
        _intro_line = shape_text_for_arabic("This report provides a comprehensive overview of the bank's control management system as of")
        _key_metrics = shape_text_for_arabic('Key Metrics:')
        _total_controls = shape_text_for_arabic('Total Controls')
        _unmapped_controls = shape_text_for_arabic('Unmapped Controls')
        _pending_preparer = shape_text_for_arabic('Pending Preparer')
        _pending_checker = shape_text_for_arabic('Pending Checker')
        _pending_reviewer = shape_text_for_arabic('Pending Reviewer')
        _pending_acceptance = shape_text_for_arabic('Pending Acceptance')
        summary_text = f"""
        {_intro_line} {datetime.now().strftime('%B %d, %Y')}.
        
        <b>{_key_metrics}</b><br/>
        • {_total_controls}: {controls_data.get('totalControls', 0):,}<br/>
        • {_unmapped_controls}: {controls_data.get('unmappedControls', 0):,}<br/>
        • {_pending_preparer}: {controls_data.get('pendingPreparer', 0):,}<br/>
        • {_pending_checker}: {controls_data.get('pendingChecker', 0):,}<br/>
        • {_pending_reviewer}: {controls_data.get('pendingReviewer', 0):,}<br/>
        • {_pending_acceptance}: {controls_data.get('pendingAcceptance', 0):,}<br/>
        """
        content.append(Paragraph(summary_text, styles['Normal']))
        content.append(Spacer(1, 20))
    
    # Section flags
    show_dept = header_config.get('showDepartmentChart', True)
    show_risk = header_config.get('showRiskChart', True)
    show_overall = header_config.get('showOverallTable', True)

    # Controls by Department Table
    if show_dept and 'departmentDistribution' in controls_data and controls_data['departmentDistribution']:
        content.append(Paragraph("Controls by Department", styles['Heading2']))
        
        # Create table data
        table_data = [[shape_text_for_arabic('Department'), shape_text_for_arabic('Controls Count')]]
        for dept in controls_data['departmentDistribution']:
            dep_name = shape_text_for_arabic(str(dept.get('name', 'N/A')))
            dep_cnt = shape_text_for_arabic(str(dept.get('value', 0)))
            table_data.append([dep_name, dep_cnt])
        
        # Create table
        table = Table(table_data, colWidths=[4*inch, 1.5*inch])
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), table_header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), ARABIC_FONT_NAME or 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), table_body_bg),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]
        if ARABIC_FONT_NAME:
            table_style.append(('FONTNAME', (0, 1), (-1, -1), ARABIC_FONT_NAME))
        table.setStyle(TableStyle(table_style))
        
        content.append(table)
        content.append(Spacer(1, 20))

        # Add chart if available
        try:
            plt.figure(figsize=(8, 4))
            dept_names = [d.get('name', 'N/A') for d in controls_data['departmentDistribution']]
            dept_counts = [d.get('value', 0) for d in controls_data['departmentDistribution']]
            plt.bar(dept_names, dept_counts)
            plt.title('Controls by Department')
            plt.xlabel('Department')
            plt.ylabel('Count')
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            # Save chart to buffer
            chart_buffer = io.BytesIO()
            plt.savefig(chart_buffer, format='png', dpi=150, bbox_inches='tight')
            chart_buffer.seek(0)
            chart_img = Image(chart_buffer)
            chart_img.drawWidth = 6*inch
            chart_img.drawHeight = 3*inch
            content.append(chart_img)
            content.append(Spacer(1, 20))
            plt.close()
        except Exception as e:
            print(f"Chart generation error: {e}")
            pass
    
    # Controls by Risk Response Table
    if show_risk and 'statusDistribution' in controls_data and controls_data['statusDistribution']:
        content.append(Paragraph("Controls by Risk Response", styles['Heading2']))
        r_table_data = [[shape_text_for_arabic('Risk Response'), shape_text_for_arabic('Count')]]
        for r in controls_data['statusDistribution']:
            rr = shape_text_for_arabic(str(r.get('name', 'N/A')))
            cnt = shape_text_for_arabic(str(r.get('value', 0)))
            r_table_data.append([rr, cnt])
        r_table = Table(r_table_data, colWidths=[4*inch, 1.5*inch])
        r_style = [
            ('BACKGROUND', (0, 0), (-1, 0), table_header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), ARABIC_FONT_NAME or 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), table_body_bg),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]
        if ARABIC_FONT_NAME:
            r_style.append(('FONTNAME', (0, 1), (-1, -1), ARABIC_FONT_NAME))
        r_table.setStyle(TableStyle(r_style))
        content.append(r_table)
        content.append(Spacer(1, 20))

        # Add pie chart if available
        try:
            plt.figure(figsize=(8, 4))
            risk_responses = [r.get('name', 'N/A') for r in controls_data['statusDistribution']]
            risk_counts = [r.get('value', 0) for r in controls_data['statusDistribution']]
            plt.pie(risk_counts, labels=risk_responses, autopct='%1.1f%%')
            plt.title('Controls by Risk Response')
            plt.tight_layout()
            
            # Save chart to buffer
            chart_buffer = io.BytesIO()
            plt.savefig(chart_buffer, format='png', dpi=150, bbox_inches='tight')
            chart_buffer.seek(0)
            chart_img = Image(chart_buffer)
            chart_img.drawWidth = 6*inch
            chart_img.drawHeight = 3*inch
            content.append(chart_img)
            content.append(Spacer(1, 20))
            plt.close()
        except Exception as e:
            print(f"Pie chart generation error: {e}")
            pass

    # Card-only table (when scoped card export)
    if header_config.get('renderOnly') == 'card' and controls_data.get('cardRows') is not None:
        title_txt = header_config.get('cardTitle') or 'Card Details'
        content.append(Paragraph(shape_text_for_arabic(title_txt), styles['Heading2']))
        card_table_data = [[shape_text_for_arabic('Index'), shape_text_for_arabic('Code'), shape_text_for_arabic('Control Name')]]
        for idx, row in enumerate(controls_data['cardRows'], 1):
            idx_str = str(idx)
            cd = shape_text_for_arabic(str(row.get('control_code') or row.get('Control Code') or 'N/A'))
            nm = shape_text_for_arabic(str(row.get('control_name') or row.get('Control Name') or 'N/A'))
            card_table_data.append([idx_str, cd, nm])
        
        # Convert control names to Paragraphs for multi-line wrapping
        body_rows = []
        for r_idx, row in enumerate(card_table_data):
            if r_idx == 0:
                body_rows.append(row)
                continue
            cell_style = ParagraphStyle('Cell', parent=styles['Normal'], wordWrap='CJK', leading=10)
            if ARABIC_FONT_NAME:
                cell_style.fontName = ARABIC_FONT_NAME
            name_para = Paragraph(row[2], cell_style)  # Control name in column 2
            body_rows.append([row[0], row[1], name_para])  # Index, Code, Name
        
        c_table = Table(body_rows, colWidths=[0.8*inch, 1.2*inch, doc.width - 2*inch])
        c_style = [
            ('BACKGROUND', (0, 0), (-1, 0), table_header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), ARABIC_FONT_NAME or 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), table_body_bg),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]
        if ARABIC_FONT_NAME:
            c_style.append(('FONTNAME', (0, 1), (-1, -1), ARABIC_FONT_NAME))
        c_table.setStyle(TableStyle(c_style))
        content.append(c_table)
        content.append(Spacer(1, 20))

    # Overall Statuses Table (all rows, no limit)
    if show_overall and 'statusOverview' in controls_data and controls_data['statusOverview']:
        content.append(Paragraph(shape_text_for_arabic("Control Status Overview"), styles['Heading2']))
        
        # Create table data
        table_data = [[shape_text_for_arabic('Index'), shape_text_for_arabic('Code'), shape_text_for_arabic('Control Name'), shape_text_for_arabic('Preparer'), shape_text_for_arabic('Checker'), shape_text_for_arabic('Reviewer'), shape_text_for_arabic('Acceptance')]]
        for idx, control in enumerate(controls_data['statusOverview'], 1):  # All rows with index
            # Use correct field names from statusOverview data
            idx_str = str(idx)
            code = control.get('code', 'N/A')  # Use actual code field
            name_full = control.get('name', 'N/A')
            # Status fields are objects with 'value' property
            preparer = control.get('preparerStatus', {}).get('value', 'N/A') if isinstance(control.get('preparerStatus'), dict) else control.get('preparerStatus', 'N/A')
            checker = control.get('checkerStatus', {}).get('value', 'N/A') if isinstance(control.get('checkerStatus'), dict) else control.get('checkerStatus', 'N/A')
            reviewer = control.get('reviewerStatus', {}).get('value', 'N/A') if isinstance(control.get('reviewerStatus'), dict) else control.get('reviewerStatus', 'N/A')
            acceptance = control.get('acceptanceStatus', {}).get('value', 'N/A') if isinstance(control.get('acceptanceStatus'), dict) else control.get('acceptanceStatus', 'N/A')
            row = [idx_str, code, name_full, preparer, checker, reviewer, acceptance]
            table_data.append([shape_text_for_arabic(str(x)) for x in row])
        
        # Create table with flexible first column and wrapped paragraphs
        # Convert the first column cells to Paragraphs to enable wrapping and multi-line rows
        body_rows = []
        for r_idx, row in enumerate(table_data):
            if r_idx == 0:
                body_rows.append(row)
                continue
            cell_style = ParagraphStyle('Cell', parent=styles['Normal'], wordWrap='CJK', leading=12)
            if ARABIC_FONT_NAME:
                cell_style.fontName = ARABIC_FONT_NAME
                cell_style.alignment = 2  # Right-align for Arabic
            name_para = Paragraph(row[2], cell_style)  # Control name in column 2
            body_rows.append([row[0], row[1], name_para, row[3], row[4], row[5], row[6]])  # Index, Code, Name, Preparer, Checker, Reviewer, Acceptance
        # Make first column occupy remaining width; others fixed
        name_col_w = max(2.5*inch, doc.width - 5*inch)
        table = Table(body_rows, colWidths=[0.5*inch, 1*inch, name_col_w, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
        status_style = [
            ('BACKGROUND', (0, 0), (-1, 0), table_header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), ARABIC_FONT_NAME or 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), table_body_bg),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8)
        ]
        if ARABIC_FONT_NAME:
            status_style.append(('FONTNAME', (0, 1), (-1, -1), ARABIC_FONT_NAME))
        table.setStyle(TableStyle(status_style))
        
        content.append(table)
        content.append(Spacer(1, 20))
    
    # Add footer
    footer_items = []
    if header_config.get('footerShowDate', True):
        footer_items.append(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}")
    if header_config.get('footerShowConfidentiality', True):
        footer_items.append(str(header_config.get('footerConfidentialityText', BANK_CONFIG['report_footer'])))

    if footer_items:
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
        if ARABIC_FONT_NAME:
            footer_style.fontName = ARABIC_FONT_NAME
        align = str(header_config.get('footerAlign', 'center')).lower()
        if align == 'left':
            footer_style.alignment = 0
        elif align == 'right':
            footer_style.alignment = 2
        else:
            footer_style.alignment = 1
        for line in footer_items:
            content.append(Paragraph(shape_text_for_arabic(line), footer_style))

    # Watermark callbacks (draw after content using onPageEnd)
    def _draw_watermark(canv, _doc):
        try:
            if not header_config.get('watermarkEnabled', False):
                return
            wm_text = str(header_config.get('watermarkText', 'CONFIDENTIAL'))
            wm_text = shape_text_for_arabic(wm_text)
            canv.saveState()
            # opacity via fill color approximation
            opacity = max(0.05, min(0.3, (header_config.get('watermarkOpacity', 10) or 10) / 100.0))
            gray = 0.6 + (0.4 * (1 - opacity))
            canv.setFillColorRGB(gray, gray, gray)
            font_name = ARABIC_FONT_NAME if ARABIC_FONT_NAME else 'Helvetica'
            canv.setFont(font_name, 48)
            page_width, page_height = A4
            canv.translate(page_width / 2.0, page_height / 2.0)
            if header_config.get('watermarkDiagonal', True):
                canv.rotate(45)
            canv.drawCentredString(0, 0, wm_text)
            canv.restoreState()
        except Exception:
            pass

    # Frames and template to allow onPageEnd watermark
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
    template = PageTemplate(id='watermarked', frames=[frame], onPageEnd=_draw_watermark)
    doc.addPageTemplates([template])

    # Build PDF
    doc.build(content)
    buffer.seek(0)
    return buffer.getvalue()

def generate_controls_excel_report(controls_data: dict, start_date: str = None, end_date: str = None, header_config: dict = None, chart_type: str = None):
    """Generate professional Excel report for Controls Dashboard"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Helpers for colors
    header_hex = str((header_config or {}).get('tableHeaderBgColor', '#1F4E79')).replace('#','')
    body_hex = str((header_config or {}).get('tableBodyBgColor', '#FFFFFF')).replace('#','')
    if len(header_hex) not in (6,8):
        header_hex = '1F4E79'
    if len(body_hex) not in (6,8):
        body_hex = 'FFFFFF'

    # Check if this is a scoped export
    is_scoped = header_config.get('scoped', False)
    render_only = header_config.get('renderOnly', None)
    
    # Debug logging
    print(f"Excel Debug - is_scoped: {is_scoped}, render_only: {render_only}")
    print(f"Excel Debug - onlyOverallTable: {header_config.get('onlyOverallTable', False)}")
    print(f"Excel Debug - onlyCard: {header_config.get('onlyCard', False)}")
    print(f"Excel Debug - cardType: {header_config.get('cardType', None)}")
    print(f"Excel Debug - controls_data keys: {list(controls_data.keys()) if controls_data else 'None'}")

    # Create Summary Sheet (only for full reports, not for scoped exports)
    if not is_scoped:
        print("Excel Debug - Creating Executive Summary sheet")
        summary_ws = wb.create_sheet("Executive Summary")
    
    # Bank header if enabled (only for Executive Summary)
    if not is_scoped and header_config and header_config.get('includeHeader', True):
        if header_config.get('showLogo', True) and header_config.get('logoBase64') and XLImage and PILImage:
            try:
                # Decode base64 and save to a temporary in-memory file
                img_bytes = base64.b64decode(header_config['logoBase64'].split(',')[-1])
                img_buf = io.BytesIO(img_bytes)
                pil_img = PILImage.open(img_buf)
                # Constrain size
                desired_h = int(header_config.get('logoHeight', 36) or 36)
                if desired_h > 64:
                    desired_h = 64
                w, h = pil_img.size
                if h > 0:
                    scale = desired_h / h
                    new_w = int(w * scale)
                    max_w = 180  # px cap for Excel
                    if new_w > max_w:
                        scale = max_w / w
                        desired_h = int(h * scale)
                        new_w = max_w
                    pil_img = pil_img.resize((new_w, desired_h))
                # Save to buffer in PNG format
                out_buf = io.BytesIO()
                pil_img.save(out_buf, format='PNG')
                out_buf.seek(0)
                xl_image = XLImage(out_buf)
                # Place logo according to position (A1 for left, merge cells for centering)
                pos = str(header_config.get('logoPosition', 'left')).lower()
                if pos == 'center':
                    summary_ws.merge_cells('A1:D3')
                    summary_ws.add_image(xl_image, 'B1')
                elif pos == 'right':
                    summary_ws.add_image(xl_image, 'E1')
                else:
                    summary_ws.add_image(xl_image, 'A1')
            except Exception:
                # Fallback to text header
                summary_ws['A1'] = f"{BANK_CONFIG['logo']} {BANK_CONFIG['name']}"
                summary_ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
        else:
            # Text fallback header
            summary_ws['A1'] = f"{BANK_CONFIG['logo']} {BANK_CONFIG['name']}"
            summary_ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
        # Bank info rows (respect toggle)
        if header_config.get('showBankInfo', True):
            summary_ws['A2'] = header_config.get('bankAddress', BANK_CONFIG['address'])
            summary_ws['A2'].font = Font(size=10, color="666666")
            summary_ws['A3'] = f"Tel: {header_config.get('bankPhone', BANK_CONFIG['phone'])} | Web: {header_config.get('bankWebsite', BANK_CONFIG['website'])}"
            summary_ws['A3'].font = Font(size=10, color="666666")
        
        # Report title with custom configuration
        title_text = header_config.get('title', 'Controls Dashboard Report')
        subtitle = header_config.get('subtitle', '')
        
        if subtitle:
            title_text += f" - {subtitle}"
        
        if start_date and end_date:
            title_text += f" ({start_date} to {end_date})"
        elif start_date:
            title_text += f" (From {start_date})"
        
        summary_ws['A5'] = title_text
        summary_ws['A5'].font = Font(size=14, bold=True, color="1F4E79")
        
        if header_config.get('showDate', True):
            summary_ws['A6'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            summary_ws['A6'].font = Font(size=10, color="666666")
    
    # Key metrics (only for Executive Summary)
    if not is_scoped:
        summary_ws['A8'] = "Key Metrics"
        summary_ws['A8'].font = Font(size=12, bold=True)
        
        metrics = [
            ("Total Controls", controls_data.get('totalControls', 0)),
            ("Unmapped Controls", controls_data.get('unmappedControls', 0)),
            ("Pending Preparer", controls_data.get('pendingPreparer', 0)),
            ("Pending Checker", controls_data.get('pendingChecker', 0)),
            ("Pending Reviewer", controls_data.get('pendingReviewer', 0)),
            ("Pending Acceptance", controls_data.get('pendingAcceptance', 0))
        ]
        
        for i, (metric, value) in enumerate(metrics, 9):
            summary_ws[f'A{i}'] = metric
            summary_ws[f'B{i}'] = value
            summary_ws[f'B{i}'].number_format = '#,##0'
    
    # Section flags
    show_dept = header_config.get('showDepartmentChart', True)
    show_risk = header_config.get('showRiskChart', True)
    show_overall = header_config.get('showOverallTable', True)

    # Controls by Department Sheet
    if (show_dept and 'departmentDistribution' in controls_data and controls_data['departmentDistribution'] and 
        (not is_scoped or render_only == 'department' or (render_only == 'chart' and chart_type == 'department'))):
        print("Excel Debug - Creating Department sheet")
        dept_ws = wb.create_sheet("Controls by Department")
        
        # Add header to department sheet if enabled and scoped
        if is_scoped and header_config and header_config.get('includeHeader', True):
            # Add logo if available
            if header_config.get('showLogo', True) and header_config.get('logoBase64') and XLImage and PILImage:
                try:
                    img_bytes = base64.b64decode(header_config['logoBase64'].split(',')[-1])
                    img_buf = io.BytesIO(img_bytes)
                    pil_img = PILImage.open(img_buf)
                    desired_h = int(header_config.get('logoHeight', 36) or 36)
                    if desired_h > 64:
                        desired_h = 64
                    w, h = pil_img.size
                    if h > 0:
                        scale = desired_h / h
                        new_w = int(w * scale)
                        max_w = 180
                        if new_w > max_w:
                            scale = max_w / w
                            desired_h = int(h * scale)
                            new_w = max_w
                        pil_img = pil_img.resize((new_w, desired_h))
                    out_buf = io.BytesIO()
                    pil_img.save(out_buf, format='PNG')
                    out_buf.seek(0)
                    xl_image = XLImage(out_buf)
                    pos = str(header_config.get('logoPosition', 'left')).lower()
                    if pos == 'center':
                        dept_ws.merge_cells('A1:D3')
                        dept_ws.add_image(xl_image, 'B1')
                    elif pos == 'right':
                        dept_ws.add_image(xl_image, 'E1')
                    else:
                        dept_ws.add_image(xl_image, 'A1')
                except Exception:
                    pass
            
            # Add title
            title_text = header_config.get('title', 'Controls by Department Analysis')
            subtitle = header_config.get('subtitle', '')
            if subtitle:
                title_text += f" - {subtitle}"
            
            if start_date and end_date:
                title_text += f" ({start_date} to {end_date})"
            elif start_date:
                title_text += f" (From {start_date})"
            
            dept_ws['A5'] = title_text
            dept_ws['A5'].font = Font(size=14, bold=True, color="1F4E79")
            
            if header_config.get('showDate', True):
                dept_ws['A6'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
                dept_ws['A6'].font = Font(size=10, color="666666")
            
            # Start data from row 8
            data_start_row = 8
        else:
            data_start_row = 1
        
        # Headers
        dept_ws[f'A{data_start_row}'] = "Department"
        dept_ws[f'B{data_start_row}'] = "Controls Count"
        dept_ws[f'A{data_start_row}'].font = Font(bold=True, color="FFFFFF")
        dept_ws[f'B{data_start_row}'].font = Font(bold=True, color="FFFFFF")
        dept_ws[f'A{data_start_row}'].fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
        dept_ws[f'B{data_start_row}'].fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
        
        # Data
        for i, dept in enumerate(controls_data['departmentDistribution'], data_start_row + 1):
            dept_ws[f'A{i}'] = dept.get('name', 'N/A')
            dept_ws[f'B{i}'] = dept.get('value', 0)
            dept_ws[f'B{i}'].number_format = '#,##0'
            dept_ws[f'A{i}'].fill = PatternFill(start_color=body_hex, end_color=body_hex, fill_type="solid")
            dept_ws[f'B{i}'].fill = PatternFill(start_color=body_hex, end_color=body_hex, fill_type="solid")
        
        # Add chart
        chart = BarChart()
        chart.title = "Controls by Department"
        chart.x_axis.title = "Department"
        chart.y_axis.title = "Count"
        
        data = Reference(dept_ws, min_col=2, min_row=data_start_row, max_row=data_start_row + len(controls_data['departmentDistribution']))
        cats = Reference(dept_ws, min_col=1, min_row=data_start_row + 1, max_row=data_start_row + len(controls_data['departmentDistribution']))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        dept_ws.add_chart(chart, f"D{data_start_row + 2}")
    
    # Controls by Risk Response Sheet
    if (show_risk and 'statusDistribution' in controls_data and controls_data['statusDistribution'] and 
        (not is_scoped or render_only == 'risk' or (render_only == 'chart' and chart_type == 'risk'))):
        print("Excel Debug - Creating Risk Response sheet")
        risk_ws = wb.create_sheet("Controls by Risk Response")
        
        # Add header to risk sheet if enabled and scoped
        if is_scoped and header_config and header_config.get('includeHeader', True):
            # Add logo if available
            if header_config.get('showLogo', True) and header_config.get('logoBase64') and XLImage and PILImage:
                try:
                    img_bytes = base64.b64decode(header_config['logoBase64'].split(',')[-1])
                    img_buf = io.BytesIO(img_bytes)
                    pil_img = PILImage.open(img_buf)
                    desired_h = int(header_config.get('logoHeight', 36) or 36)
                    if desired_h > 64:
                        desired_h = 64
                    w, h = pil_img.size
                    if h > 0:
                        scale = desired_h / h
                        new_w = int(w * scale)
                        max_w = 180
                        if new_w > max_w:
                            scale = max_w / w
                            desired_h = int(h * scale)
                            new_w = max_w
                        pil_img = pil_img.resize((new_w, desired_h))
                    out_buf = io.BytesIO()
                    pil_img.save(out_buf, format='PNG')
                    out_buf.seek(0)
                    xl_image = XLImage(out_buf)
                    pos = str(header_config.get('logoPosition', 'left')).lower()
                    if pos == 'center':
                        risk_ws.merge_cells('A1:D3')
                        risk_ws.add_image(xl_image, 'B1')
                    elif pos == 'right':
                        risk_ws.add_image(xl_image, 'E1')
                    else:
                        risk_ws.add_image(xl_image, 'A1')
                except Exception:
                    pass
            
            # Add title
            title_text = header_config.get('title', 'Controls by Risk Response Analysis')
            subtitle = header_config.get('subtitle', '')
            if subtitle:
                title_text += f" - {subtitle}"
            
            if start_date and end_date:
                title_text += f" ({start_date} to {end_date})"
            elif start_date:
                title_text += f" (From {start_date})"
            
            risk_ws['A5'] = title_text
            risk_ws['A5'].font = Font(size=14, bold=True, color="1F4E79")
            
            if header_config.get('showDate', True):
                risk_ws['A6'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
                risk_ws['A6'].font = Font(size=10, color="666666")
            
            # Start data from row 8
            data_start_row = 8
        else:
            data_start_row = 1
        
        # Headers
        risk_ws[f'A{data_start_row}'] = "Risk Response"
        risk_ws[f'B{data_start_row}'] = "Controls Count"
        risk_ws[f'A{data_start_row}'].font = Font(bold=True, color="FFFFFF")
        risk_ws[f'B{data_start_row}'].font = Font(bold=True, color="FFFFFF")
        risk_ws[f'A{data_start_row}'].fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
        risk_ws[f'B{data_start_row}'].fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
        
        # Data
        for i, risk in enumerate(controls_data['statusDistribution'], data_start_row + 1):
            risk_ws[f'A{i}'] = risk.get('name', 'N/A')
            risk_ws[f'B{i}'] = risk.get('value', 0)
            risk_ws[f'B{i}'].number_format = '#,##0'
            risk_ws[f'A{i}'].fill = PatternFill(start_color=body_hex, end_color=body_hex, fill_type="solid")
            risk_ws[f'B{i}'].fill = PatternFill(start_color=body_hex, end_color=body_hex, fill_type="solid")
        
        # Add chart
        chart = BarChart()
        chart.title = "Controls by Risk Response"
        chart.x_axis.title = "Risk Response"
        chart.y_axis.title = "Count"
        
        data = Reference(risk_ws, min_col=2, min_row=data_start_row, max_row=data_start_row + len(controls_data['statusDistribution']))
        cats = Reference(risk_ws, min_col=1, min_row=data_start_row + 1, max_row=data_start_row + len(controls_data['statusDistribution']))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        risk_ws.add_chart(chart, f"D{data_start_row + 2}")
    
    # Overall Statuses Sheet - Always create if data exists and it's requested
    should_create_overall = (
        'statusOverview' in controls_data and
        controls_data['statusOverview'] and
        (
            not is_scoped or 
            render_only == 'overall' or 
            header_config.get('onlyOverallTable', False) or
            show_overall
        )
    )
    
    if should_create_overall:
        print("Excel Debug - Creating Overall Statuses sheet")
        status_ws = wb.create_sheet("Control Statuses")
        
        # Add header to status sheet if enabled and scoped
        if is_scoped and header_config and header_config.get('includeHeader', True):
            # Add logo if available
            if header_config.get('showLogo', True) and header_config.get('logoBase64') and XLImage and PILImage:
                try:
                    img_bytes = base64.b64decode(header_config['logoBase64'].split(',')[-1])
                    img_buf = io.BytesIO(img_bytes)
                    pil_img = PILImage.open(img_buf)
                    desired_h = int(header_config.get('logoHeight', 36) or 36)
                    if desired_h > 64:
                        desired_h = 64
                    w, h = pil_img.size
                    if h > 0:
                        scale = desired_h / h
                        new_w = int(w * scale)
                        max_w = 180
                        if new_w > max_w:
                            scale = max_w / w
                            desired_h = int(h * scale)
                            new_w = max_w
                        pil_img = pil_img.resize((new_w, desired_h))
                    out_buf = io.BytesIO()
                    pil_img.save(out_buf, format='PNG')
                    out_buf.seek(0)
                    xl_image = XLImage(out_buf)
                    pos = str(header_config.get('logoPosition', 'left')).lower()
                    if pos == 'center':
                        status_ws.merge_cells('A1:D3')
                        status_ws.add_image(xl_image, 'B1')
                    elif pos == 'right':
                        status_ws.add_image(xl_image, 'E1')
                    else:
                        status_ws.add_image(xl_image, 'A1')
                except Exception:
                    pass
            
            # Add title
            title_text = header_config.get('title', 'All Control Statuses Overview')
            subtitle = header_config.get('subtitle', '')
            if subtitle:
                title_text += f" - {subtitle}"
            
            if start_date and end_date:
                title_text += f" ({start_date} to {end_date})"
            elif start_date:
                title_text += f" (From {start_date})"
            
            status_ws['A5'] = title_text
            status_ws['A5'].font = Font(size=14, bold=True, color="1F4E79")
            
            if header_config.get('showDate', True):
                status_ws['A6'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
                status_ws['A6'].font = Font(size=10, color="666666")
            
            # Start data from row 8
            data_start_row = 8
        else:
            data_start_row = 1
        
        # Headers
        headers = ['Index', 'Code', 'Control Name', 'Preparer Status', 'Checker Status', 'Reviewer Status', 'Acceptance Status']
        # Colors from config
        for i, header in enumerate(headers, 1):
            cell = status_ws.cell(row=data_start_row, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data (support snake_case and title-case) - ALL ROWS, no limit
        for row_idx, control in enumerate(controls_data['statusOverview'], data_start_row + 1):
            idx = str(row_idx - data_start_row)  # Index number starting from 1
            code = control.get('code', 'N/A')  # Use actual code field
            name = control.get('name', 'N/A')
            # Status fields are objects with 'value' property
            preparer = control.get('preparerStatus', {}).get('value', 'N/A') if isinstance(control.get('preparerStatus'), dict) else control.get('preparerStatus', 'N/A')
            checker = control.get('checkerStatus', {}).get('value', 'N/A') if isinstance(control.get('checkerStatus'), dict) else control.get('checkerStatus', 'N/A')
            reviewer = control.get('reviewerStatus', {}).get('value', 'N/A') if isinstance(control.get('reviewerStatus'), dict) else control.get('reviewerStatus', 'N/A')
            acceptance = control.get('acceptanceStatus', {}).get('value', 'N/A') if isinstance(control.get('acceptanceStatus'), dict) else control.get('acceptanceStatus', 'N/A')
            status_ws.cell(row=row_idx, column=1, value=idx)
            status_ws.cell(row=row_idx, column=2, value=code)
            status_ws.cell(row=row_idx, column=3, value=name)
            status_ws.cell(row=row_idx, column=4, value=preparer)
            status_ws.cell(row=row_idx, column=5, value=checker)
            status_ws.cell(row=row_idx, column=6, value=reviewer)
            status_ws.cell(row=row_idx, column=7, value=acceptance)
            # Apply body background color
            for col in range(1, 8):
                status_ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=body_hex, end_color=body_hex, fill_type="solid")
        
        # Auto-adjust column widths
        for column in status_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            status_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Apply table colors to Executive Summary metrics block as well (only for Executive Summary)
    if not is_scoped:
        summary_ws['A8'].fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
        summary_ws['A8'].font = Font(size=12, bold=True, color="FFFFFF")
        for i, (metric, value) in enumerate([
            ("Total Controls", controls_data.get('totalControls', 0)),
            ("Unmapped Controls", controls_data.get('unmappedControls', 0)),
            ("Pending Preparer", controls_data.get('pendingPreparer', 0)),
            ("Pending Checker", controls_data.get('pendingChecker', 0)),
            ("Pending Reviewer", controls_data.get('pendingReviewer', 0)),
            ("Pending Acceptance", controls_data.get('pendingAcceptance', 0))
        ], 9):
            summary_ws[f'A{i}'].fill = PatternFill(start_color=body_hex, end_color=body_hex, fill_type="solid")
            summary_ws[f'B{i}'].fill = PatternFill(start_color=body_hex, end_color=body_hex, fill_type="solid")

    # Card-only table (when scoped card export) - Always create if data exists
    print(f"Excel Debug - cardRows in controls_data: {'cardRows' in controls_data}")
    print(f"Excel Debug - cardRows data: {controls_data.get('cardRows') is not None if 'cardRows' in controls_data else 'No cardRows key'}")
    print(f"Excel Debug - renderOnly: {header_config.get('renderOnly')}")
    print(f"Excel Debug - is_scoped: {is_scoped}")
    print(f"Excel Debug - onlyCard: {header_config.get('onlyCard', False)}")
    
    should_create_card = (
        'cardRows' in controls_data and 
        controls_data['cardRows'] is not None and 
        (
            header_config.get('renderOnly') == 'card' or 
            is_scoped or
            header_config.get('onlyCard', False)
        )
    )
    print(f"Excel Debug - should_create_card: {should_create_card}")
    
    if should_create_card:
        print("Excel Debug - Creating Card Details sheet")
        card_ws = wb.create_sheet("Card Details")
        
        # Add header to card sheet if enabled
        if header_config and header_config.get('includeHeader', True):
            # Add logo if available
            if header_config.get('showLogo', True) and header_config.get('logoBase64') and XLImage and PILImage:
                try:
                    img_bytes = base64.b64decode(header_config['logoBase64'].split(',')[-1])
                    img_buf = io.BytesIO(img_bytes)
                    pil_img = PILImage.open(img_buf)
                    desired_h = int(header_config.get('logoHeight', 36) or 36)
                    if desired_h > 64:
                        desired_h = 64
                    w, h = pil_img.size
                    if h > 0:
                        scale = desired_h / h
                        new_w = int(w * scale)
                        max_w = 180
                        if new_w > max_w:
                            scale = max_w / w
                            desired_h = int(h * scale)
                            new_w = max_w
                        pil_img = pil_img.resize((new_w, desired_h))
                    out_buf = io.BytesIO()
                    pil_img.save(out_buf, format='PNG')
                    out_buf.seek(0)
                    xl_image = XLImage(out_buf)
                    pos = str(header_config.get('logoPosition', 'left')).lower()
                    if pos == 'center':
                        card_ws.merge_cells('A1:D3')
                        card_ws.add_image(xl_image, 'B1')
                    elif pos == 'right':
                        card_ws.add_image(xl_image, 'E1')
                    else:
                        card_ws.add_image(xl_image, 'A1')
                except Exception:
                    pass
            
            # Add title
            title_text = header_config.get('title', 'Control Details Report')
            subtitle = header_config.get('subtitle', '')
            if subtitle:
                title_text += f" - {subtitle}"
            
            if start_date and end_date:
                title_text += f" ({start_date} to {end_date})"
            elif start_date:
                title_text += f" (From {start_date})"
            
            card_ws['A5'] = title_text
            card_ws['A5'].font = Font(size=14, bold=True, color="1F4E79")
            
            if header_config.get('showDate', True):
                card_ws['A6'] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
                card_ws['A6'].font = Font(size=10, color="666666")
            
            # Start data from row 8
            data_start_row = 8
        else:
            data_start_row = 1
        
        # Headers
        headers = ['Index', 'Code', 'Control Name']
        for i, header in enumerate(headers, 1):
            cell = card_ws.cell(row=data_start_row, column=i, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        for row_idx, row in enumerate(controls_data['cardRows'], data_start_row + 1):
            idx = str(row_idx - data_start_row)  # Index number starting from 1
            code = row.get('control_code') or row.get('Control Code') or 'N/A'
            name = row.get('control_name') or row.get('Control Name') or 'N/A'
            card_ws.cell(row=row_idx, column=1, value=idx)
            card_ws.cell(row=row_idx, column=2, value=code)
            card_ws.cell(row=row_idx, column=3, value=name)
            # Apply body background color
            for col in range(1, 4):
                card_ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=body_hex, end_color=body_hex, fill_type="solid")
        
        # Auto-adjust column widths
        for column in card_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            card_ws.column_dimensions[column_letter].width = adjusted_width

    # Debug: Check sheets created
    print(f"Excel Debug - Sheets created: {[ws.title for ws in wb.worksheets]}")
    
    # Ensure at least one sheet exists - Create based on available data
    if len(wb.worksheets) == 0:
        print("Excel Debug - No sheets created, creating fallback sheet with available data")
        
        # Try to create a meaningful sheet based on available data
        if controls_data and 'statusOverview' in controls_data and controls_data['statusOverview']:
            print("Excel Debug - Creating fallback sheet with overallStatuses data")
            fallback_ws = wb.create_sheet("Control Statuses")
            
            # Add headers
            headers = ['Index', 'Code', 'Control Name', 'Preparer Status', 'Checker Status', 'Reviewer Status', 'Acceptance Status']
            for i, header in enumerate(headers, 1):
                cell = fallback_ws.cell(row=1, column=i, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Add data
            for row_idx, control in enumerate(controls_data['statusOverview'], 2):
                idx = str(row_idx - 1)  # Index starting from 1 for fallback
                code = control.get('code', 'N/A')  # Use actual code field
                name = control.get('name', 'N/A')
                # Status fields are objects with 'value' property
                preparer = control.get('preparerStatus', {}).get('value', 'N/A') if isinstance(control.get('preparerStatus'), dict) else control.get('preparerStatus', 'N/A')
                checker = control.get('checkerStatus', {}).get('value', 'N/A') if isinstance(control.get('checkerStatus'), dict) else control.get('checkerStatus', 'N/A')
                reviewer = control.get('reviewerStatus', {}).get('value', 'N/A') if isinstance(control.get('reviewerStatus'), dict) else control.get('reviewerStatus', 'N/A')
                acceptance = control.get('acceptanceStatus', {}).get('value', 'N/A') if isinstance(control.get('acceptanceStatus'), dict) else control.get('acceptanceStatus', 'N/A')
                
                fallback_ws.cell(row=row_idx, column=1, value=idx)
                fallback_ws.cell(row=row_idx, column=2, value=code)
                fallback_ws.cell(row=row_idx, column=3, value=name)
                fallback_ws.cell(row=row_idx, column=4, value=preparer)
                fallback_ws.cell(row=row_idx, column=5, value=checker)
                fallback_ws.cell(row=row_idx, column=6, value=reviewer)
                fallback_ws.cell(row=row_idx, column=7, value=acceptance)
                
                # Apply body background color
                for col in range(1, 8):
                    fallback_ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=body_hex, end_color=body_hex, fill_type="solid")
                    
        elif controls_data and 'statusDistribution' in controls_data and controls_data['statusDistribution']:
            print("Excel Debug - Creating fallback sheet with Risk Response data")
            fallback_ws = wb.create_sheet("Controls by Risk Response")
            
            # Add headers
            headers = ['Risk Response', 'Controls Count']
            for i, header in enumerate(headers, 1):
                cell = fallback_ws.cell(row=1, column=i, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Add data
            for row_idx, risk in enumerate(controls_data['statusDistribution'], 2):
                risk_response = risk.get('name', 'N/A')
                count = risk.get('value', 0)
                
                fallback_ws.cell(row=row_idx, column=1, value=risk_response)
                fallback_ws.cell(row=row_idx, column=2, value=count)
                fallback_ws.cell(row=row_idx, column=2).number_format = '#,##0'
                
                # Apply body background color
                for col in range(1, 3):
                    fallback_ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=body_hex, end_color=body_hex, fill_type="solid")
                    
        elif controls_data and 'cardRows' in controls_data and controls_data['cardRows']:
            print("Excel Debug - Creating fallback sheet with cardRows data")
            fallback_ws = wb.create_sheet("Card Details")
            
            # Add headers
            headers = ['Index', 'Code', 'Control Name']
            for i, header in enumerate(headers, 1):
                cell = fallback_ws.cell(row=1, column=i, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Add data
            for row_idx, row in enumerate(controls_data['cardRows'], 2):
                idx = str(row_idx - 1)  # Index starting from 1
                code = row.get('control_code') or row.get('Control Code') or 'N/A'
                name = row.get('control_name') or row.get('Control Name') or 'N/A'
                
                fallback_ws.cell(row=row_idx, column=1, value=idx)
                fallback_ws.cell(row=row_idx, column=2, value=code)
                fallback_ws.cell(row=row_idx, column=3, value=name)
                
                # Apply body background color
                for col in range(1, 4):
                    fallback_ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=body_hex, end_color=body_hex, fill_type="solid")
        else:
            print("Excel Debug - Creating empty fallback sheet")
            fallback_ws = wb.create_sheet("Data")
            fallback_ws['A1'] = "No Data Available"
            fallback_ws['A1'].font = Font(size=14, bold=True)
            fallback_ws['A2'] = "The requested data could not be generated."
            fallback_ws['A2'].font = Font(size=12)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

@app.get("/api/grc/controls/export-pdf")
async def export_controls_pdf(
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    headerConfig: Optional[str] = None,
    includeHeader: bool = True,
    icon: str = "building",
    title: str = "Controls Dashboard Report",
    subtitle: str = "Governance, Risk & Compliance Management",
    location: str = "top",
    showLogo: bool = True,
    showDate: bool = True,
    showPageNumbers: bool = True,
    logoPosition: str = "left",
    logoBase64: Optional[str] = None,
    fontSize: str = "medium",
    fontColor: str = "#1F4E79",
    backgroundColor: str = "#FFFFFF",
    borderStyle: str = "solid",
    borderColor: str = "#E5E7EB",
    borderWidth: int = 1,
    padding: int = 20,
    margin: int = 10,
    logoHeight: int = 36,
    # Bank info
    showBankInfo: bool = True,
    bankName: str = 'PIANAT.AI',
    bankAddress: str = 'King Abdulaziz Road, Riyadh, Saudi Arabia',
    bankPhone: str = '+966 11 402 9000',
    bankWebsite: str = 'www.alrajhibank.com.sa',
    # Footer
    footerShowDate: bool = True,
    footerShowConfidentiality: bool = True,
    footerConfidentialityText: str = 'Confidential Banking Report - Internal Use Only',
    footerShowPageNumbers: bool = True,
    footerAlign: str = 'center',
    # Watermark
    watermarkEnabled: bool = False,
    watermarkText: str = 'CONFIDENTIAL',
    watermarkOpacity: int = 10,
    watermarkDiagonal: bool = True,
    # Table colors
    tableHeaderBgColor: str = '#1F4E79',
    tableBodyBgColor: str = '#FFFFFF',
    # Card scoping
    cardType: Optional[str] = None,
    onlyCard: bool = False,
    # Chart/table scoping
    onlyChart: bool = False,
    chartType: Optional[str] = None,
    onlyOverallTable: bool = False,
    onlyFunctionTable: bool = False
):
    """Export Controls Dashboard as PDF"""
    try:
        # Get controls data from Node.js backend with date filtering
        import httpx
        async with httpx.AsyncClient(timeout=60.0) as client:
            params = {}
            if startDate:
                params["startDate"] = startDate
            if endDate:
                params["endDate"] = endDate
                
            response = await client.get("http://localhost:3002/api/grc/controls", params=params)
            controls_data = response.json()
        
        # Scope controls_data if onlyCard and cardType provided
        # Build scoping flags and optionally load SQL-backed data for card tables
        section_flags = {
            'showDepartmentChart': False,
            'showRiskChart': False,
            'showOverallTable': False,
            'scoped': False,
            'renderOnly': None,
            'cardTitle': None
        }

        scope_card_rows = None

        if onlyCard and cardType:
            section_flags['scoped'] = True
            section_flags['renderOnly'] = 'card'
            # Fetch specific card endpoint from Node (uses SQL under the hood)
            endpoint_map = {
                'totalControls': 'total',
                'unmappedControls': 'unmapped',
                'pendingPreparer': 'pending-preparer',
                'pendingChecker': 'pending-checker',
                'pendingReviewer': 'pending-reviewer',
                'pendingAcceptance': 'pending-acceptance',
            }
            path = endpoint_map.get(cardType)
            if not path:
                path = 'total'
            import httpx
            async with httpx.AsyncClient(timeout=60.0) as client:
                params_cards = {}
                if startDate:
                    params_cards['startDate'] = startDate
                if endDate:
                    params_cards['endDate'] = endDate
                params_cards['page'] = 1
                params_cards['limit'] = 20000
                resp = await client.get(f"http://localhost:3002/api/grc/controls/{path}", params=params_cards)
                data = resp.json()
                scope_card_rows = data.get('data', [])
            section_flags['cardTitle'] = cardType.replace('Controls',' Controls').replace('pending','Pending ').title()
        elif onlyChart and chartType:
            # For chart exports, show only the specific chart
            section_flags['showDepartmentChart'] = False
            section_flags['showRiskChart'] = False
            section_flags['showOverallTable'] = False
            section_flags['scoped'] = True
            if chartType == 'department':
                section_flags['showDepartmentChart'] = True
                section_flags['renderOnly'] = 'department'
            elif chartType == 'risk' or chartType == 'status':
                section_flags['showRiskChart'] = True
                section_flags['renderOnly'] = 'risk'
        elif onlyOverallTable:
            # For table exports, show only the specific table
            section_flags['showDepartmentChart'] = False
            section_flags['showRiskChart'] = False
            section_flags['showOverallTable'] = True
            section_flags['scoped'] = True
            section_flags['renderOnly'] = 'overall'
        elif onlyFunctionTable:
            # For function table exports, show only the function table
            section_flags['showDepartmentChart'] = False
            section_flags['showRiskChart'] = False
            section_flags['showOverallTable'] = False
            section_flags['showControlsByFunction'] = True
            section_flags['scoped'] = True
            section_flags['renderOnly'] = 'function'
        else:
            # full report
            section_flags['showDepartmentChart'] = True
            section_flags['showRiskChart'] = True
            section_flags['showOverallTable'] = True

        # Generate PDF with custom header
        header_config = {
            'includeHeader': includeHeader,
            'icon': icon,
            'title': title,
            'subtitle': subtitle,
            'location': location,
            'showLogo': showLogo,
            'showDate': showDate,
            'showPageNumbers': showPageNumbers,
            'logoPosition': logoPosition,
            'logoBase64': logoBase64,
            'fontSize': fontSize,
            'fontColor': fontColor,
            'backgroundColor': backgroundColor,
            'borderStyle': borderStyle,
            'borderColor': borderColor,
            'borderWidth': borderWidth,
            'padding': padding,
            'margin': margin,
            'logoHeight': logoHeight,
            'showBankInfo': showBankInfo,
            'bankName': bankName,
            'bankAddress': bankAddress,
            'bankPhone': bankPhone,
            'bankWebsite': bankWebsite,
            'footerShowDate': footerShowDate,
            'footerShowConfidentiality': footerShowConfidentiality,
            'footerConfidentialityText': footerConfidentialityText,
            'footerShowPageNumbers': footerShowPageNumbers,
            'footerAlign': footerAlign,
            'watermarkEnabled': watermarkEnabled,
            'watermarkText': watermarkText,
            'watermarkOpacity': watermarkOpacity,
            'watermarkDiagonal': watermarkDiagonal,
            'tableHeaderBgColor': tableHeaderBgColor,
            'tableBodyBgColor': tableBodyBgColor,
            **section_flags
        }
        # Merge JSON headerConfig if provided (modal values take precedence)
        if headerConfig:
            try:
                import json
                parsed = json.loads(headerConfig)
                if isinstance(parsed, dict):
                    header_config.update(parsed)
            except Exception:
                # Ignore malformed JSON but continue with defaults
                pass
        # Attach scoped card rows into controls_data if present
        if scope_card_rows is not None:
            # For individual exports, preserve the full dashboard data and add card-specific data
            controls_data['cardRows'] = scope_card_rows

        # Validate data before generating PDF
        if not controls_data:
            raise ValueError("No data available for PDF generation")
        
        print(f"Generating PDF with data keys: {list(controls_data.keys())}")
        try:
            pdf_content = generate_controls_pdf_report(controls_data, startDate, endDate, header_config)
            if not pdf_content:
                raise ValueError("PDF generation returned empty content")
        except Exception as pdf_error:
            print(f"PDF generation failed: {str(pdf_error)}")
            import traceback
            traceback.print_exc()
            raise ValueError(f"PDF generation failed: {str(pdf_error)}")
        
        # Generate dynamic filename based on scope
        if onlyCard and cardType:
            card_names = {
                'totalControls': 'total-controls',
                'unmappedControls': 'unmapped-controls',
                'pendingPreparer': 'pending-preparer-controls',
                'pendingChecker': 'pending-checker-controls',
                'pendingReviewer': 'pending-reviewer-controls',
                'pendingAcceptance': 'pending-acceptance-controls'
            }
            base_name = card_names.get(cardType, 'control-list')
        elif onlyChart and chartType:
            if chartType == 'department':
                base_name = 'controls-by-department'
            elif chartType == 'risk':
                base_name = 'controls-by-risk-response'
            else:
                base_name = 'chart-analysis'
        elif onlyOverallTable:
            base_name = 'all-control-statuses'
        elif onlyFunctionTable:
            base_name = 'controls-by-function'
        else:
            base_name = 'all-data-dashboard'
        
        # Add date range if provided
        if startDate and endDate:
            base_name += f"_{startDate}_to_{endDate}"
        elif startDate:
            base_name += f"_from_{startDate}"
        
        filename = f"{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = reports_dir / filename
        
        with open(filepath, "wb") as f:
            f.write(pdf_content)
        
        return FileResponse(
            path=str(filepath),
            filename=filename,
            media_type="application/pdf"
        )
    except Exception as e:
        import traceback
        error_details = str(e) if str(e) else "Unknown error"
        traceback_str = traceback.format_exc()
        print(f"PDF Generation Error: {error_details}")
        print(f"Traceback: {traceback_str}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {error_details}")

@app.get("/api/grc/controls/test-debug")
async def test_debug():
    """Test debug endpoint"""
    import os
    debug_file = os.path.join(os.getcwd(), 'excel_debug.log')
    with open(debug_file, 'a') as f:
        f.write("Test debug endpoint called\n")
    return {"message": "Debug test successful"}

@app.get("/api/grc/controls/export-excel")
async def export_controls_excel(
    startDate: Optional[str] = None,
    endDate: Optional[str] = None,
    headerConfig: Optional[str] = None,
    includeHeader: bool = True,
    icon: str = "building",
    title: str = "Controls Dashboard Report",
    subtitle: str = "Governance, Risk & Compliance Management",
    location: str = "top",
    showLogo: bool = True,
    showDate: bool = True,
    showPageNumbers: bool = True,
    logoPosition: str = "left",
    logoBase64: Optional[str] = None,
    fontSize: str = "medium",
    fontColor: str = "#1F4E79",
    backgroundColor: str = "#FFFFFF",
    borderStyle: str = "solid",
    borderColor: str = "#E5E7EB",
    borderWidth: int = 1,
    padding: int = 20,
    margin: int = 10,
    logoHeight: int = 36,
    # Bank info
    showBankInfo: bool = True,
    bankName: str = 'PIANAT.AI',
    bankAddress: str = 'King Abdulaziz Road, Riyadh, Saudi Arabia',
    bankPhone: str = '+966 11 402 9000',
    bankWebsite: str = 'www.alrajhibank.com.sa',
    # Footer
    footerShowDate: bool = True,
    footerShowConfidentiality: bool = True,
    footerConfidentialityText: str = 'Confidential Banking Report - Internal Use Only',
    footerShowPageNumbers: bool = True,
    footerAlign: str = 'center',
    # Watermark (Excel: not rendered as overlay, but can add in header/footer as text)
    watermarkEnabled: bool = False,
    watermarkText: str = 'CONFIDENTIAL',
    # Table colors
    tableHeaderBgColor: str = '#1F4E79',
    tableBodyBgColor: str = '#FFFFFF',
    # Card scoping
    cardType: Optional[str] = None,
    onlyCard: bool = False,
    # Chart/table scoping
    onlyChart: bool = False,
    chartType: Optional[str] = None,
    onlyOverallTable: bool = False,
    onlyFunctionTable: bool = False
):
    """Export Controls Dashboard as Excel"""
    # Write debug info to file immediately
    import os
    debug_file = os.path.join(os.getcwd(), 'excel_debug.log')
    with open(debug_file, 'a') as f:
        f.write(f"Excel Debug - Function called with onlyCard: {onlyCard}, cardType: {cardType}\n")
        f.write(f"Excel Debug - Current working directory: {os.getcwd()}\n")
        f.write(f"Excel Debug - Debug file path: {debug_file}\n")
    
    try:
        # Write debug info to file
        import os
        debug_file = os.path.join(os.getcwd(), 'excel_debug.log')
        with open(debug_file, 'a') as f:
            f.write(f"Excel Debug - Starting export with onlyCard: {onlyCard} (type: {type(onlyCard)}), cardType: {cardType} (type: {type(cardType)})\n")
            f.write(f"Excel Debug - Function called with parameters: onlyCard={onlyCard}, cardType={cardType}\n")
        
        # Convert string parameters to proper types
        if isinstance(onlyCard, str):
            onlyCard = onlyCard.lower() in ['true', '1', 'yes', 'on']
        elif onlyCard is None:
            onlyCard = False
            
        if isinstance(cardType, str) and cardType.lower() in ['none', 'null', '']:
            cardType = None
            
        with open(debug_file, 'a') as f:
            f.write(f"Excel Debug - After conversion: onlyCard={onlyCard} (type: {type(onlyCard)}), cardType={cardType} (type: {type(cardType)})\n")
            f.flush()
        
        # Get controls data from Node.js backend with date filtering
        import httpx
        try:
            with open(debug_file, 'a') as f:
                f.write(f"Excel Debug - About to fetch controls data from Node.js backend\n")
                f.flush()
            
            async with httpx.AsyncClient(timeout=60.0) as client:
                params = {}
                if startDate:
                    params["startDate"] = startDate
                if endDate:
                    params["endDate"] = endDate
                
                with open(debug_file, 'a') as f:
                    f.write(f"Excel Debug - Making request to http://localhost:3002/api/grc/controls with params: {params}\n")
                    f.flush()
                
                response = await client.get("http://localhost:3002/api/grc/controls", params=params)
                
                with open(debug_file, 'a') as f:
                    f.write(f"Excel Debug - Got response with status: {response.status_code}\n")
                    f.flush()
                
                controls_data = response.json()
            
            with open(debug_file, 'a') as f:
                f.write(f"Excel Debug - Fetched controls data: {list(controls_data.keys())}\n")
                f.flush()
        except Exception as e:
            with open(debug_file, 'a') as f:
                f.write(f"Excel Debug - Error fetching controls data: {e}\n")
                f.flush()
            raise
        
        # Build header configuration
        header_config = {
            'includeHeader': includeHeader,
            'icon': icon,
            'title': title,
            'subtitle': subtitle,
            'location': location,
            'showLogo': showLogo,
            'showDate': showDate,
            'showPageNumbers': showPageNumbers,
            'logoPosition': logoPosition,
            'logoBase64': logoBase64,
            'fontSize': fontSize,
            'fontColor': fontColor,
            'backgroundColor': backgroundColor,
            'borderStyle': borderStyle,
            'borderColor': borderColor,
            'borderWidth': borderWidth,
            'padding': padding,
            'margin': margin,
            'logoHeight': logoHeight,
            'showBankInfo': showBankInfo,
            'bankName': bankName,
            'bankAddress': bankAddress,
            'bankPhone': bankPhone,
            'bankWebsite': bankWebsite,
            'footerShowDate': footerShowDate,
            'footerShowConfidentiality': footerShowConfidentiality,
            'footerConfidentialityText': footerConfidentialityText,
            'footerShowPageNumbers': footerShowPageNumbers,
            'footerAlign': footerAlign,
            'watermarkEnabled': watermarkEnabled,
            'watermarkText': watermarkText,
            'watermarkOpacity': 0.1,
            'watermarkDiagonal': True,
            'excelFreezeTopRow': True,
            'excelAutoFitColumns': True,
            'excelZebraStripes': True,
            'excelFitToWidth': True,
            'tableHeaderBgColor': tableHeaderBgColor,
            'tableBodyBgColor': tableBodyBgColor,
            'onlyCard': onlyCard,
            'cardType': cardType,
            'onlyChart': onlyChart,
            'chartType': chartType,
            'onlyOverallTable': onlyOverallTable,
            'scoped': onlyCard or onlyChart or onlyOverallTable,
            'renderOnly': 'card' if onlyCard else ('chart' if onlyChart else ('overall' if onlyOverallTable else None))
        }
        # Merge JSON headerConfig if provided (modal values take precedence)
        if headerConfig:
            try:
                import json
                parsed = json.loads(headerConfig)
                if isinstance(parsed, dict):
                    header_config.update(parsed)
            except Exception:
                pass
        
        # Get scoped card data if needed
        scope_card_rows = None
        with open(debug_file, 'a') as f:
            f.write(f"Excel Debug - Checking card data fetching: onlyCard={onlyCard}, cardType={cardType}\n")
            f.write(f"Excel Debug - Condition check: onlyCard={onlyCard}, cardType={cardType}, onlyCard and cardType={onlyCard and cardType}\n")
            f.flush()  # Force write to file
        
        if onlyCard and cardType:
            with open(debug_file, 'a') as f:
                f.write(f"Excel Debug - Entering card data fetching section\n")
            # Map frontend card types to backend API card types
            card_type_mapping = {
                'totalControls': 'total',
                'unmappedControls': 'unmapped',
                'pendingPreparer': 'pending-preparer',
                'pendingChecker': 'pending-checker',
                'pendingReviewer': 'pending-reviewer',
                'pendingAcceptance': 'pending-acceptance'
            }
            backend_card_type = card_type_mapping.get(cardType, cardType)
            with open(debug_file, 'a') as f:
                f.write(f"Excel Debug - Fetching card data for cardType: {cardType} -> {backend_card_type}\n")
            try:
                card_api_url = f"http://localhost:3002/api/grc/controls/{backend_card_type}"
                with open(debug_file, 'a') as f:
                    f.write(f"Excel Debug - Card API URL: {card_api_url}\n")
                
                # Add page and limit parameters for card data fetching
                card_params = {**params, 'page': 1, 'limit': 10000}  # Fetch all data
                
                async with httpx.AsyncClient(timeout=60.0) as client:
                    card_response = await client.get(card_api_url, params=card_params)
                    with open(debug_file, 'a') as f:
                        f.write(f"Excel Debug - Card API Response Status: {card_response.status_code}\n")
                    card_data = card_response.json()
                    scope_card_rows = card_data.get('data', [])
                    with open(debug_file, 'a') as f:
                        f.write(f"Excel Debug - Card data length: {len(scope_card_rows)}\n")
            except Exception as e:
                with open(debug_file, 'a') as f:
                    f.write(f"Error fetching card data: {e}\n")
                import traceback
                traceback.print_exc()
                scope_card_rows = []
        
        # Add section flags - for individual exports, include all data
        if onlyChart and chartType:
            # For chart exports, show only the specific chart
            section_flags = {
                'showSummary': False,
                'showDepartmentChart': False,
                'showStatusChart': False,
                'showOverallTable': False,
                'showControlsByFunction': False,
                'scoped': True
            }
            if chartType == 'department':
                section_flags['showDepartmentChart'] = True
            elif chartType == 'risk' or chartType == 'status':
                section_flags['showStatusChart'] = True
        elif onlyOverallTable:
            # For table exports, show only the specific table
            section_flags = {
                'showSummary': False,
                'showDepartmentChart': False,
                'showStatusChart': False,
                'showOverallTable': True,
                'showControlsByFunction': False,
                'scoped': True
            }
        elif onlyFunctionTable:
            # For function table exports, show only the function table
            section_flags = {
                'showSummary': False,
                'showDepartmentChart': False,
                'showStatusChart': False,
                'showOverallTable': False,
                'showControlsByFunction': True,
                'scoped': True
            }
        else:
            # Full report or card exports
            section_flags = {
                'showSummary': not onlyChart and not onlyOverallTable,
                'showDepartmentChart': not onlyCard and not onlyOverallTable,
                'showStatusChart': not onlyCard and not onlyOverallTable,
                'showOverallTable': not onlyCard and not onlyChart,
                'showControlsByFunction': not onlyCard and not onlyChart and not onlyOverallTable
            }
        
        # Merge section flags into header config
        header_config = {
            **header_config,
            **section_flags
        }
        # Attach scoped card rows into controls_data if present
        if scope_card_rows is not None:
            # For individual exports, preserve the full dashboard data and add card-specific data
            controls_data['cardRows'] = scope_card_rows
            with open(debug_file, 'a') as f:
                f.write(f"Excel Debug - Added cardRows to controls_data: {len(scope_card_rows)} rows\n")
                f.write(f"Excel Debug - First card row: {scope_card_rows[0] if scope_card_rows else 'No data'}\n")
                f.write(f"Excel Debug - controls_data keys: {list(controls_data.keys())}\n")
                f.flush()
        
        with open(debug_file, 'a') as f:
            f.write(f"Excel Debug - Before Excel generation: cardRows in controls_data = {'cardRows' in controls_data}\n")
            if 'cardRows' in controls_data:
                f.write(f"Excel Debug - cardRows length: {len(controls_data.get('cardRows', []))}\n")
            f.flush()
        
        excel_content = generate_controls_excel_report(controls_data, startDate, endDate, header_config, chartType)
        
        # Generate dynamic filename based on scope
        if onlyCard and cardType:
            card_names = {
                'totalControls': 'total-controls',
                'unmappedControls': 'unmapped-controls',
                'pendingPreparer': 'pending-preparer-controls',
                'pendingChecker': 'pending-checker-controls',
                'pendingReviewer': 'pending-reviewer-controls',
                'pendingAcceptance': 'pending-acceptance-controls'
            }
            filename = f"grc-{card_names.get(cardType, cardType)}-export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif onlyChart and chartType:
            chart_names = {
                'department': 'controls-by-department',
                'status': 'controls-by-status'
            }
            filename = f"grc-{chart_names.get(chartType, chartType)}-chart-export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif onlyOverallTable:
            filename = f"grc-overall-control-statuses-export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif onlyFunctionTable:
            filename = f"grc-controls-by-function-export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = f"grc-controls-dashboard-export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Save to file
        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)
        filepath = reports_dir / filename
        
        with open(filepath, "wb") as f:
            f.write(excel_content)
    
        return FileResponse(
            path=str(filepath),
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        import traceback
        error_details = str(e) if str(e) else "Unknown error"
        traceback_str = traceback.format_exc()
        print(f"Excel Generation Error: {error_details}")
        print(f"Traceback: {traceback_str}")
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {error_details}")

@app.get("/api/reports/generate-excel")
async def generate_excel_report(
    report_type: str = "summary",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Generate Excel report with charts"""
    # Get data
    data = {
        "summary": {"total": 100, "active": 80, "inactive": 20},
        "charts": [
            {"name": "Chart 1", "value": 50},
            {"name": "Chart 2", "value": 30}
        ]
    }
    
    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    # Add data
    ws['A1'] = "Report Type"
    ws['B1'] = report_type
    ws['A2'] = "Total"
    ws['B2'] = data['summary']['total']
    ws['A3'] = "Active"
    ws['B3'] = data['summary']['active']
    ws['A4'] = "Inactive"
    ws['B4'] = data['summary']['inactive']
    
    # Save to file
    reports_dir = Path("reports")
    reports_dir.mkdir(exist_ok=True)
    filename = f"report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = reports_dir / filename
    
    wb.save(filepath)
    
    return FileResponse(
        path=str(filepath),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
