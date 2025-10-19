"""
Reusable export utilities for dashboard reports
"""
import io
import json
from datetime import datetime
from typing import Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Default bank configuration
DEFAULT_BANK_CONFIG = {
    'name': 'PIANAT.AI',
    'address': 'King Abdulaziz Road, Riyadh, Saudi Arabia',
    'phone': '+966 11 402 9000',
    'website': 'www.pianat.ai',
    'logo': '🏦'
}

def get_default_header_config(dashboard_type: str = "dashboard") -> Dict[str, Any]:
    """Get default header configuration based on dashboard type"""
    defaults = {
        "risks": {
            "includeHeader": True,
            "icon": "shield-exclamation",
            "title": "Risks Dashboard Report",
            "subtitle": "Risk Management & Assessment Monitoring",
            "fontColor": "#1F4E79",
            "watermarkEnabled": True,
            "tableHeaderBgColor": "#DC2626"
        },
        "controls": {
            "includeHeader": True,
            "icon": "building",
            "title": "Controls Dashboard Report", 
            "subtitle": "Comprehensive Analysis Report",
            "fontColor": "#1F4E79",
            "watermarkEnabled": True
        },
        "incidents": {
            "includeHeader": True,
            "icon": "exclamation-triangle",
            "title": "Incidents Dashboard Report", 
            "subtitle": "Comprehensive Analysis Report",
            "fontColor": "#1F4E79",
            "watermarkEnabled": True
        },
        "kris": {
            "includeHeader": True,
            "icon": "chart-bar",
            "title": "KRIs Dashboard Report", 
            "subtitle": "Key Risk Indicators Monitoring & Assessment",
            "fontColor": "#1F4E79",
            "watermarkEnabled": True
        }
    }
    
    base_config = {
        "location": "top",
        "showLogo": True,
        "showDate": True,
        "showPageNumbers": True,
        "logoPosition": "left",
        "logoFile": None,
        "logoBase64": None,
        "fontSize": "medium",
        "backgroundColor": "#FFFFFF",
        "borderStyle": "solid",
        "borderColor": "#E5E7EB",
        "borderWidth": 1,
        "padding": 20,
        "margin": 10,
        "logoHeight": 36,
        "showBankInfo": True,
        "bankName": DEFAULT_BANK_CONFIG['name'],
        "bankAddress": DEFAULT_BANK_CONFIG['address'],
        "bankPhone": DEFAULT_BANK_CONFIG['phone'],
        "bankWebsite": DEFAULT_BANK_CONFIG['website'],
        "footerShowDate": True,
        "footerShowConfidentiality": True,
        "footerConfidentialityText": "Confidential Report - Internal Use Only",
        "footerShowPageNumbers": True,
        "footerAlign": "center",
        "watermarkEnabled": False,
        "watermarkText": "CONFIDENTIAL",
        "watermarkOpacity": 10,
        "watermarkDiagonal": True,
        "excelFreezeTopRow": True,
        "excelAutoFitColumns": True,
        "excelZebraStripes": True,
        "excelFitToWidth": True,
        "tableHeaderBgColor": "#1F4E79",
        "tableBodyBgColor": "#FFFFFF"
    }
    
    # Merge with dashboard-specific defaults
    dashboard_defaults = defaults.get(dashboard_type, defaults["controls"])
    base_config.update(dashboard_defaults)
    
    return base_config

def merge_header_config(user_config: Optional[Dict[str, Any]], dashboard_type: str = "dashboard") -> Dict[str, Any]:
    """Merge user configuration with defaults"""
    default_config = get_default_header_config(dashboard_type)
    
    if not user_config:
        return default_config
    
    # Handle string configuration (from URL parameters)
    if isinstance(user_config, str):
        try:
            user_config = json.loads(user_config)
        except json.JSONDecodeError:
            return default_config
    
    if not isinstance(user_config, dict):
        return default_config
    
    # Merge configurations
    merged_config = default_config.copy()
    merged_config.update(user_config)
    
    return merged_config

def add_watermark_to_pdf(story, header_config: Dict[str, Any]) -> None:
    """Add watermark to PDF story as background text"""
    if not header_config.get('watermarkEnabled', False):
        return
    
    watermark_text = header_config.get('watermarkText', 'CONFIDENTIAL')
    
    try:
        from reportlab.platypus import Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        # Create watermark style
        watermark_style = ParagraphStyle(
            'Watermark',
            parent=getSampleStyleSheet()['Normal'],
            fontSize=48,
            textColor=colors.lightgrey,
            alignment=1,  # Center alignment
            spaceAfter=0,
            spaceBefore=0,
            leading=60
        )
        
        # Add watermark as a paragraph (will appear behind content)
        watermark_para = Paragraph(watermark_text, watermark_style)
        
        # Insert watermark at the beginning of the story
        story.insert(0, watermark_para)
        story.insert(1, Spacer(1, 0.1*inch))  # Small spacer
        
        pass
        
    except Exception as e:
        pass

def add_watermark_to_excel_sheet(worksheet, header_config: Dict[str, Any]) -> None:
    """Add watermark to Excel worksheet as background image"""
    if not header_config.get('watermarkEnabled', False):
        return
    
    watermark_text = header_config.get('watermarkText', 'CONFIDENTIAL')
    
    try:
        from PIL import Image, ImageDraw, ImageFont
        import io
        
        # Create a transparent watermark image
        width, height = 400, 200
        watermark_img = Image.new('RGBA', (width, height), (255, 255, 255, 0))  # Transparent background
        draw = ImageDraw.Draw(watermark_img)
        
        # Try to use a system font, fallback to default
        try:
            font = ImageFont.truetype("arial.ttf", 24)
        except:
            try:
                font = ImageFont.truetype("Arial.ttf", 24)
            except:
                font = ImageFont.load_default()
        
        # Get text size
        bbox = draw.textbbox((0, 0), watermark_text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        
        # Center the text
        x = (width - text_width) // 2
        y = (height - text_height) // 2
        
        # Draw text with transparency
        opacity = header_config.get('watermarkOpacity', 10) / 100.0
        text_color = (128, 128, 128, int(255 * opacity))  # Gray with opacity
        
        draw.text((x, y), watermark_text, font=font, fill=text_color)
        
        # Rotate the image if diagonal watermark is enabled
        if header_config.get('watermarkDiagonal', True):
            watermark_img = watermark_img.rotate(45, expand=True)
        
        # Save to BytesIO
        watermark_buf = io.BytesIO()
        watermark_img.save(watermark_buf, format='PNG')
        watermark_buf.seek(0)
        
        # Add as background image to worksheet
        from openpyxl.drawing.image import Image as XLImage
        xl_watermark = XLImage(watermark_buf)
        
        # Position watermark in center of worksheet
        worksheet.add_image(xl_watermark, 'H10')  # Center-ish position
        
        pass
        
    except Exception as e:
        pass
        # Fallback: Add watermark text to header
        try:
            worksheet.HeaderFooter.center_header.text = f"&C{watermark_text}"
            pass
        except Exception as header_error:
            pass

def add_bank_info_to_excel_sheet(worksheet, header_config: Dict[str, Any], start_row: int = 1) -> int:
    """Add bank information to Excel worksheet and return the next available row"""
    if not header_config.get('showBankInfo', True):
        return start_row
    
    bank_name = header_config.get('bankName', DEFAULT_BANK_CONFIG['name'])
    bank_address = header_config.get('bankAddress', DEFAULT_BANK_CONFIG['address'])
    bank_phone = header_config.get('bankPhone', DEFAULT_BANK_CONFIG['phone'])
    bank_website = header_config.get('bankWebsite', DEFAULT_BANK_CONFIG['website'])
    
    pass
    
    # Add logo if available
    logo_added = False
    if header_config.get('showLogo', True) and header_config.get('logoBase64'):
        try:
            import base64
            import io
            from PIL import Image as PILImage
            from openpyxl.drawing.image import Image as XLImage
            
            logo_base64 = header_config['logoBase64']
            logo_position = header_config.get('logoPosition', 'left').lower()
            
            # Decode base64 and create image
            img_bytes = base64.b64decode(logo_base64.split(',')[-1])
            img_buf = io.BytesIO(img_bytes)
            pil_img = PILImage.open(img_buf)
            
            # Constrain size
            desired_h = int(header_config.get('logoHeight', 36) or 36)
            if desired_h > 64:
                desired_h = 64
            w, h = pil_img.size
            if h > 0:
                scale = desired_h / h
                new_w = int(w * scale)
                max_w = 180
                if new_w > max_w:
                    new_w = max_w
                    scale = new_w / w
                    desired_h = int(h * scale)
                pil_img = pil_img.resize((new_w, desired_h), PILImage.Resampling.LANCZOS)
            
            # Save to BytesIO for openpyxl
            logo_buf = io.BytesIO()
            pil_img.save(logo_buf, format='PNG')
            logo_buf.seek(0)
            
            # Create openpyxl image
            xl_image = XLImage(logo_buf)
            
            # Position logo based on logoPosition
            if logo_position == 'left':
                worksheet.add_image(xl_image, 'A1')
            elif logo_position == 'center':
                # Center the logo (approximate)
                worksheet.add_image(xl_image, 'C1')
            elif logo_position == 'right':
                # Right align the logo (approximate)
                worksheet.add_image(xl_image, 'E1')
            
            logo_added = True
            pass
            
        except Exception as e:
            pass
            logo_added = False
    
    # Adjust start row if logo was added
    if logo_added:
        start_row += 2  # Leave space for logo
    
    # Add bank information
    worksheet.cell(row=start_row, column=1, value=bank_name)
    worksheet.cell(row=start_row + 1, column=1, value=bank_address)
    worksheet.cell(row=start_row + 2, column=1, value=f"Tel: {bank_phone} | Web: {bank_website}")
    
    # Style bank information
    font_color = header_config.get('fontColor', '#1F4E79')
    # Convert hex color to RGB for openpyxl
    if font_color.startswith('#'):
        font_color = font_color[1:]  # Remove #
    
    worksheet.cell(row=start_row, column=1).font = Font(size=12, bold=True, color=font_color)
    worksheet.cell(row=start_row + 1, column=1).font = Font(size=10, color=font_color)
    worksheet.cell(row=start_row + 2, column=1).font = Font(size=10, color=font_color)
    
    pass
    
    return start_row + 3  # Return next available row

def add_date_to_excel_sheet(worksheet, header_config: Dict[str, Any], row: int, column: int = 1) -> int:
    """Add generation date to Excel worksheet"""
    if not header_config.get('showDate', True):
        return row
    
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    worksheet.cell(row=row, column=column, value=f"Generated on: {current_date}")
    worksheet.cell(row=row, column=column).font = Font(size=10, italic=True)
    
    pass
    
    return row + 1

def create_excel_summary_sheet(workbook: Workbook, sheet_name: str, data: Dict[str, Any], 
                              header_config: Dict[str, Any], dashboard_type: str = "dashboard") -> None:
    """Create a summary sheet with bank info and data"""
    worksheet = workbook.create_sheet(sheet_name)
    
    # Add bank information
    current_row = add_bank_info_to_excel_sheet(worksheet, header_config, 1)
    
    # Add generation date
    current_row = add_date_to_excel_sheet(worksheet, header_config, current_row)
    
    # Add empty row for spacing
    current_row += 1
    
    # Add summary data based on dashboard type
    if dashboard_type == "risks":
        summary_data = [
            ['Metric', 'Value'],
            ['Total Risks', data.get('totalRisks', 0)],
            ['High Risks', len([r for r in data.get('inherentVsResidual', []) if (r.get('inherent_value', 0) or 0) >= 7])],
            ['Medium Risks', len([r for r in data.get('inherentVsResidual', []) if 4 <= (r.get('inherent_value', 0) or 0) < 7])],
            ['Low Risks', len([r for r in data.get('inherentVsResidual', []) if (r.get('inherent_value', 0) or 0) < 4])],
        ]
    else:  # controls
        summary_data = [
            ['Metric', 'Value'],
            ['Total Controls', data.get('total', 0)],
            ['Pending Preparer', data.get('pendingPreparer', 0)],
            ['Pending Checker', data.get('pendingChecker', 0)],
            ['Pending Reviewer', data.get('pendingReviewer', 0)],
            ['Pending Acceptance', data.get('pendingAcceptance', 0)],
        ]
    
    # Add summary table
    for i, row_data in enumerate(summary_data):
        for j, cell_value in enumerate(row_data):
            worksheet.cell(row=current_row + i, column=j+1, value=cell_value)
    
    # Style summary table
    font_color = header_config.get('fontColor', '#1F4E79').replace('#', '')
    header_bg_color = header_config.get('tableHeaderBgColor', '#1F4E79').replace('#', '')
    body_bg_color = header_config.get('tableBodyBgColor', '#FFFFFF').replace('#', '')
    
    for row in range(current_row, current_row + len(summary_data)):
        for col in range(1, len(summary_data[0]) + 1):
            cell = worksheet.cell(row=row, column=col)
            if row == current_row:  # Header row
                # Remove # if present for Excel compatibility
                header_color = header_bg_color
                if header_color.startswith('#'):
                    header_color = header_color[1:]
                cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            else:
                # Remove # if present for Excel compatibility
                body_color = body_bg_color
                if body_color.startswith('#'):
                    body_color = body_color[1:]
                cell.fill = PatternFill(start_color=body_color, end_color=body_color, fill_type='solid')
    
    pass

def auto_fit_excel_columns(worksheet) -> None:
    """Auto-fit column widths in Excel worksheet"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def create_pdf_header_elements(header_config: Dict[str, Any]) -> list:
    """Create PDF header elements from configuration"""
    elements = []
    styles = getSampleStyleSheet()
    
    # Create custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor(header_config.get('fontColor', '#1F4E79')),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor(header_config.get('fontColor', '#1F4E79')),
        alignment=TA_CENTER,
        spaceAfter=6
    )
    
    bank_style = ParagraphStyle(
        'BankInfo',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor(header_config.get('fontColor', '#1F4E79')),
        alignment=TA_CENTER,
        spaceAfter=3
    )
    
    # Add title and subtitle
    if header_config.get('includeHeader', True):
        elements.append(Paragraph(header_config.get('title', 'Dashboard Report'), title_style))
        elements.append(Paragraph(header_config.get('subtitle', 'Comprehensive Analysis Report'), subtitle_style))
        
        # Add bank information
        if header_config.get('showBankInfo', True):
            bank_name = header_config.get('bankName', DEFAULT_BANK_CONFIG['name'])
            bank_address = header_config.get('bankAddress', DEFAULT_BANK_CONFIG['address'])
            bank_phone = header_config.get('bankPhone', DEFAULT_BANK_CONFIG['phone'])
            bank_website = header_config.get('bankWebsite', DEFAULT_BANK_CONFIG['website'])
            
            elements.append(Paragraph(f"{DEFAULT_BANK_CONFIG['logo']} {bank_name}", bank_style))
            elements.append(Paragraph(bank_address, bank_style))
            elements.append(Paragraph(f"Tel: {bank_phone} | Web: {bank_website}", bank_style))
        
        # Add generation date
        if header_config.get('showDate', True):
            current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            elements.append(Paragraph(f"Generated on: {current_date}", bank_style))
        
        elements.append(Spacer(1, 20))
    
    return elements

def create_pdf_footer_elements(header_config: Dict[str, Any]) -> list:
    """Create PDF footer elements from configuration"""
    elements = []
    styles = getSampleStyleSheet()
    
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    if header_config.get('footerShowDate', True):
        current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elements.append(Paragraph(f"Generated on: {current_date}", footer_style))
    
    if header_config.get('footerShowConfidentiality', True):
        confidentiality_text = header_config.get('footerConfidentialityText', 'Confidential Report - Internal Use Only')
        elements.append(Paragraph(confidentiality_text, footer_style))
    
    return elements
