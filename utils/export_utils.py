"""
Reusable export utilities for dashboard reports
CENTRALIZED CONFIGURATION SYSTEM - Single Source of Truth for PDF & Excel
All default values are defined here, users can override via header_config parameter
"""
import io
import json
from datetime import datetime
from typing import Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ==============================================================================
# CENTRAL CONFIGURATION SYSTEM
# ==============================================================================

# Default bank information
DEFAULT_BANK_CONFIG = {
    'name': 'PIANAT.AI',
    'address': 'Bank address',
    'phone': 'Bank phone',
    'website': 'www.website.com',
    'logo': '🏦'
}

def get_complete_default_config() -> Dict[str, Any]:
    """
    Get complete default configuration for PDF and Excel exports
    This is the SINGLE SOURCE OF TRUTH for all default values
    """
    return {
        # ========== HEADER SECTION ==========
        "includeHeader": True,              # Show/hide entire header section
        "showLogo": True,                  # Show/hide logo
        "showDate": True,                   # Show generation date in header
        "showPageNumbers": True,            # Show page numbers
        "showBankInfo": True,               # Show/hide bank information
        "bankInfoLocation": "top",          # Bank info position: top/bottom/none
        
        # ========== TITLE & SUBTITLE ==========
        "title": "Dashboard Report",        # Main title
        "subtitle": "",                     # Subtitle (optional)
        "icon": "chart-line",               # Icon identifier
        
        # ========== LOGO CONFIGURATION ==========
        "logoPosition": "left",             # Logo position: left/center/right
        "logoHeight": 36,                   # Logo height in points
        "logoFile": None,                   # Path to logo file
        "logoBase64": None,                 # Base64 encoded logo string
        
        # ========== BANK INFORMATION ==========
        "bankName": DEFAULT_BANK_CONFIG['name'],
        "bankAddress": DEFAULT_BANK_CONFIG['address'],
        "bankPhone": DEFAULT_BANK_CONFIG['phone'],
        "bankWebsite": DEFAULT_BANK_CONFIG['website'],
        
        # ========== COLOR SCHEME ==========
        "fontColor": "#1F4E79",             # Main font color (dark blue)
        "backgroundColor": "#FFFFFF",       # Background color (white)
        
        # ========== TABLE STYLING (PDF & Excel) ==========
        "tableHeaderBgColor": "#1F4E79",    # Table header background (dark blue)
        "tableHeaderTextColor": "#FFFFFF",  # Table header text color (white)
        "tableBodyBgColor": "#FFFFFF",      # Table body background (white)
        "tableBodyTextColor": "#000000",    # Table body text color (black)
        
        # ========== BORDER & BORDERS ==========
        "borderStyle": "solid",            # Border style
        "borderColor": "#C9CED6",           # Border color (light gray to match risks look)
        "borderWidth": 1,                   # Border width in points
        
        # ========== MARGINS & SPACING (Defaults tuned for dense tables) ==========
        "leftMargin": 24,                  # ~0.33 inch
        "rightMargin": 24,                 # ~0.33 inch
        "topMargin": 36,                   # ~0.5 inch
        "bottomMargin": 36,                # ~0.5 inch
        "padding": 12,                     # Slightly tighter cell padding
        "margin": 8,                       # General spacing
        
        # ========== FONTS ==========
        "fontSize": "medium",              # Font size: small/medium/large
        
        # ========== FOOTER SECTION ==========
        "footerShowDate": True,            # Show date in footer
        "footerShowConfidentiality": True, # Show confidentiality notice
        "footerConfidentialityText": "Confidential Report - Internal Use Only",
        "footerShowPageNumbers": True,    # Show page numbers in footer
        "footerAlign": "center",          # Footer alignment: left/center/right
        
        # ========== WATERMARK SECTION ==========
        "watermarkEnabled": False,         # Enable watermark
        "watermarkText": "CONFIDENTIAL",   # Watermark text
        "watermarkOpacity": 10,            # Watermark opacity (0-100)
        "watermarkDiagonal": True,         # Diagonal watermark placement
        
        # ========== CHART SECTION ==========
        "chart_type": "bar",               # Chart type: bar/pie/line
        "chart_data": None,                # Chart data: {labels: [], values: []}
        
        # ========== EXCEL SPECIFIC SETTINGS ==========
        "excelFreezeTopRow": True,         # Freeze header row
        "excelAutoFitColumns": True,       # Auto-fit column widths
        "excelZebraStripes": True,         # Alternating row colors
        "excelFitToWidth": True,          # Fit to page width
        
        # ========== PDF SPECIFIC SETTINGS ==========
        "pdfPageOrientation": "portrait",  # Page orientation: portrait/landscape

        # ========== TABLE BEHAVIOR DEFAULTS ==========
        "maxCellChars": 300,               # Truncate overly long cell text with ellipsis
        "maxRowsPerTable": 25,             # Split very large tables into batches
        
        # ========== LOCATION & LAYOUT ==========
        "location": "top",                 # Main content location
    }

def get_dashboard_specific_overrides() -> Dict[str, Dict[str, Any]]:
    """
    Dashboard-specific configuration overrides
    These override the base defaults for each dashboard type
    """
    # Unify defaults across all dashboards to keep a consistent look like risks screenshot
    generic = {
        "title": "Dashboard Report",
        "subtitle": "Comprehensive Analysis Report",
        "icon": "chart-line",
        "watermarkEnabled": True,
    }
    return {
        "risks": generic.copy(),
        "controls": generic.copy(),
        "incidents": generic.copy(),
        "kris": generic.copy(),
        "dynamic": generic.copy(),
    }

def get_default_header_config(dashboard_type: str = "controls") -> Dict[str, Any]:
    """
    Get complete default header configuration for a specific dashboard type
    Combines base defaults with dashboard-specific overrides
    """
    # Start with complete default configuration
    config = get_complete_default_config()
    
    # Apply dashboard-specific overrides
    overrides = get_dashboard_specific_overrides()
    dashboard_overrides = overrides.get(dashboard_type, overrides.get("controls", {}))
    
    # Merge dashboard-specific overrides into base config
    config.update(dashboard_overrides)
    
    return config

def merge_header_config(user_config: Optional[Dict[str, Any]], dashboard_type: str = "controls") -> Dict[str, Any]:
    """
    Merge user configuration with centralized defaults
    
    Args:
        user_config: User-provided configuration (can be dict, JSON string, or None)
        dashboard_type: Type of dashboard (controls, risks, incidents, kris, dynamic)
    
    Returns:
        Complete merged configuration with all defaults applied
    """
    # Get defaults for this dashboard type
    default_config = get_default_header_config(dashboard_type)
    
    # If no user config provided, return defaults
    if not user_config:
        return default_config
    
    # Handle string configuration (from URL parameters)
    if isinstance(user_config, str):
        try:
            user_config = json.loads(user_config)
        except json.JSONDecodeError:
            # Invalid JSON string, return defaults
            return default_config
    
    # Ensure it's a dictionary
    if not isinstance(user_config, dict):
        return default_config
    
    # Deep merge user config into defaults (user values take precedence)
    merged_config = default_config.copy()
    merged_config.update(user_config)
    
    return merged_config

def add_watermark_to_pdf(story, header_config: Dict[str, Any]) -> None:
    """Add watermark to PDF story as background text positioned in the middle"""
    if not header_config.get('watermarkEnabled', False):
        return
    
    watermark_text = header_config.get('watermarkText', 'CONFIDENTIAL')
    watermark_opacity = header_config.get('watermarkOpacity', 10)
    watermark_diagonal = header_config.get('watermarkDiagonal', True)
    
    try:
        from reportlab.platypus import Paragraph, Spacer, KeepTogether
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus.flowables import Flowable
        from reportlab.graphics.shapes import Drawing, String
        from reportlab.graphics import renderPDF
        
        # Create a custom watermark flowable that positions in the middle
        class WatermarkFlowable(Flowable):
            def __init__(self, text, opacity, diagonal):
                self.text = text
                self.opacity = opacity / 100.0
                self.diagonal = diagonal
                # Required attributes for Flowable
                self.width = 0
                self.height = 0
                
            def draw(self):
                # Get canvas dimensions
                canvas = self.canv
                width = canvas._pagesize[0]
                height = canvas._pagesize[1]
                
                # Save current state
                canvas.saveState()
                
                # Set opacity
                canvas.setFillAlpha(self.opacity)
                canvas.setStrokeAlpha(self.opacity)
                
                # Set text color to light gray
                canvas.setFillColor(colors.lightgrey)
                canvas.setStrokeColor(colors.lightgrey)
                
                # Calculate center position
                center_x = width / 2
                center_y = height / 2
                
                # Set font and size
                canvas.setFont("Helvetica-Bold", 48)
                
                if self.diagonal:
                    # Rotate for diagonal effect
                    canvas.rotate(45)
                    # Adjust position for rotation
                    canvas.drawCentredString(center_x, center_y, self.text)
                else:
                    # Horizontal text
                    canvas.drawCentredString(center_x, center_y, self.text)
                
                # Restore state
                canvas.restoreState()
        
        # Create and add watermark
        watermark = WatermarkFlowable(watermark_text, watermark_opacity, watermark_diagonal)
        
        # Insert watermark at the beginning so it appears behind content
        story.insert(0, watermark)
        story.insert(1, Spacer(1, 0.1*inch))  # Small spacer
        
    except Exception as e:
        # Fallback to simple paragraph watermark
        try:
            from reportlab.platypus import Paragraph
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib import colors
            
            # Create watermark style
            watermark_style = ParagraphStyle(
                'Watermark',
                parent=getSampleStyleSheet()['Normal'],
                fontSize=48,
                textColor=colors.lightgrey,
                alignment=1,  # Center alignment
                spaceAfter=0,
                spaceBefore=0,
                leading=60
            )
            
            # Add watermark as a paragraph
            watermark_para = Paragraph(watermark_text, watermark_style)
            story.insert(0, watermark_para)
            story.insert(1, Spacer(1, 0.1*inch))
        except:
            pass

def add_watermark_to_excel_sheet(worksheet, header_config: Dict[str, Any]) -> None:
    """Add watermark to Excel worksheet as background image positioned in the middle"""
    if not header_config.get('watermarkEnabled', False):
        return
    
    watermark_text = header_config.get('watermarkText', 'CONFIDENTIAL')
    watermark_opacity = header_config.get('watermarkOpacity', 10)
    watermark_diagonal = header_config.get('watermarkDiagonal', True)
    
    try:
        from PIL import Image, ImageDraw, ImageFont
        import io
        import math
        
        # Get worksheet dimensions to calculate center position
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Calculate center position (approximate)
        center_row = max(1, max_row // 2)
        center_col = max(1, max_col // 2)
        
        # Create a larger watermark image for better visibility
        width, height = 600, 300
        watermark_img = Image.new('RGBA', (width, height), (255, 255, 255, 0))  # Transparent background
        draw = ImageDraw.Draw(watermark_img)
        
        # Try to use a system font, fallback to default
        try:
            font = ImageFont.truetype("arial.ttf", 36)
        except:
            try:
                font = ImageFont.truetype("Arial.ttf", 36)
            except:
                try:
                    font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 36)
                except:
                    font = ImageFont.load_default()
        
        # Get text size
        bbox = draw.textbbox((0, 0), watermark_text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        
        # Center the text
        x = (width - text_width) // 2
        y = (height - text_height) // 2
        
        # Draw text with transparency
        opacity = watermark_opacity / 100.0
        text_color = (128, 128, 128, int(255 * opacity))  # Gray with opacity
        
        draw.text((x, y), watermark_text, font=font, fill=text_color)
        
        # Rotate the image if diagonal watermark is enabled
        if watermark_diagonal:
            # Rotate 45 degrees for diagonal effect
            watermark_img = watermark_img.rotate(45, expand=True)
        
        # Save to BytesIO
        watermark_buf = io.BytesIO()
        watermark_img.save(watermark_buf, format='PNG')
        watermark_buf.seek(0)
        
        # Add as background image to worksheet
        from openpyxl.drawing.image import Image as XLImage
        xl_watermark = XLImage(watermark_buf)
        
        # Position watermark in center of worksheet
        # Convert center position to Excel cell reference
        from openpyxl.utils import get_column_letter
        center_cell = f"{get_column_letter(center_col)}{center_row}"
        worksheet.add_image(xl_watermark, center_cell)
        
    except Exception as e:
        # Fallback: Add watermark text to header
        try:
            if watermark_diagonal:
                # For diagonal effect, add multiple spaces to create visual effect
                diagonal_text = " ".join(watermark_text)
                worksheet.HeaderFooter.center_header.text = f"&C{diagonal_text}"
            else:
                worksheet.HeaderFooter.center_header.text = f"&C{watermark_text}"
        except Exception as header_error:
            pass

def add_bank_info_to_excel_sheet(worksheet, header_config: Dict[str, Any], start_row: int = 1) -> int:
    """Add bank information to Excel worksheet and return the next available row"""
    if not header_config.get('showBankInfo', True):
        return start_row
    
    bank_name = header_config.get('bankName', DEFAULT_BANK_CONFIG['name'])
    bank_address = header_config.get('bankAddress', DEFAULT_BANK_CONFIG['address'])
    bank_phone = header_config.get('bankPhone', DEFAULT_BANK_CONFIG['phone'])
    bank_website = header_config.get('bankWebsite', DEFAULT_BANK_CONFIG['website'])
    
    pass
    
    # Add logo if available
    logo_added = False
    if header_config.get('showLogo', True) and header_config.get('logoBase64'):
        try:
            import base64
            import io
            from PIL import Image as PILImage
            from openpyxl.drawing.image import Image as XLImage
            
            logo_base64 = header_config['logoBase64']
            logo_position = header_config.get('logoPosition', 'left').lower()
            
            # Decode base64 and create image
            img_bytes = base64.b64decode(logo_base64.split(',')[-1])
            img_buf = io.BytesIO(img_bytes)
            pil_img = PILImage.open(img_buf)
            
            # Constrain size
            desired_h = int(header_config.get('logoHeight', 36) or 36)
            if desired_h > 64:
                desired_h = 64
            w, h = pil_img.size
            if h > 0:
                scale = desired_h / h
                new_w = int(w * scale)
                max_w = 180
                if new_w > max_w:
                    new_w = max_w
                    scale = new_w / w
                    desired_h = int(h * scale)
                pil_img = pil_img.resize((new_w, desired_h), PILImage.Resampling.LANCZOS)
            
            # Save to BytesIO for openpyxl
            logo_buf = io.BytesIO()
            pil_img.save(logo_buf, format='PNG')
            logo_buf.seek(0)
            
            # Create openpyxl image
            xl_image = XLImage(logo_buf)
            
            # Position logo based on logoPosition
            if logo_position == 'left':
                worksheet.add_image(xl_image, 'A1')
            elif logo_position == 'center':
                # Center the logo (approximate)
                worksheet.add_image(xl_image, 'C1')
            elif logo_position == 'right':
                # Right align the logo (approximate)
                worksheet.add_image(xl_image, 'E1')
            
            logo_added = True
            pass
            
        except Exception as e:
            pass
            logo_added = False
    
    # Adjust start row if logo was added
    if logo_added:
        start_row += 2  # Leave space for logo
    
    # Add bank information
    worksheet.cell(row=start_row, column=1, value=bank_name)
    worksheet.cell(row=start_row + 1, column=1, value=bank_address)
    worksheet.cell(row=start_row + 2, column=1, value=f"Tel: {bank_phone} | Web: {bank_website}")
    
    # Style bank information
    font_color = header_config.get('fontColor', '#1F4E79')
    # Convert hex color to RGB for openpyxl
    if font_color.startswith('#'):
        font_color = font_color[1:]  # Remove #
    
    worksheet.cell(row=start_row, column=1).font = Font(size=12, bold=True, color=font_color)
    worksheet.cell(row=start_row + 1, column=1).font = Font(size=10, color=font_color)
    worksheet.cell(row=start_row + 2, column=1).font = Font(size=10, color=font_color)
    
    pass
    
    return start_row + 3  # Return next available row

def add_date_to_excel_sheet(worksheet, header_config: Dict[str, Any], row: int, column: int = 1) -> int:
    """Add generation date to Excel worksheet"""
    if not header_config.get('showDate', True):
        return row
    
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    worksheet.cell(row=row, column=column, value=f"Generated on: {current_date}")
    worksheet.cell(row=row, column=column).font = Font(size=10, italic=True)
    
    pass
    
    return row + 1

def create_excel_summary_sheet(workbook: Workbook, sheet_name: str, data: Dict[str, Any], 
                              header_config: Dict[str, Any], dashboard_type: str = "dashboard") -> None:
    """Create a summary sheet with bank info and data"""
    worksheet = workbook.create_sheet(sheet_name)
    
    # Add bank information
    current_row = add_bank_info_to_excel_sheet(worksheet, header_config, 1)
    
    # Add generation date
    current_row = add_date_to_excel_sheet(worksheet, header_config, current_row)
    
    # Add empty row for spacing
    current_row += 1
    
    # Add summary data based on dashboard type
    if dashboard_type == "risks":
        summary_data = [
            ['Metric', 'Value'],
            ['Total Risks', data.get('totalRisks', 0)],
            ['High Risks', len([r for r in data.get('inherentVsResidual', []) if (r.get('inherent_value', 0) or 0) >= 7])],
            ['Medium Risks', len([r for r in data.get('inherentVsResidual', []) if 4 <= (r.get('inherent_value', 0) or 0) < 7])],
            ['Low Risks', len([r for r in data.get('inherentVsResidual', []) if (r.get('inherent_value', 0) or 0) < 4])],
        ]
    else:  # controls
        summary_data = [
            ['Metric', 'Value'],
            ['Total Controls', data.get('total', 0)],
            ['Pending Preparer', data.get('pendingPreparer', 0)],
            ['Pending Checker', data.get('pendingChecker', 0)],
            ['Pending Reviewer', data.get('pendingReviewer', 0)],
            ['Pending Acceptance', data.get('pendingAcceptance', 0)],
        ]
    
    # Add summary table
    for i, row_data in enumerate(summary_data):
        for j, cell_value in enumerate(row_data):
            worksheet.cell(row=current_row + i, column=j+1, value=cell_value)
    
    # Style summary table
    font_color = header_config.get('fontColor', '#1F4E79').replace('#', '')
    header_bg_color = header_config.get('tableHeaderBgColor', '#1F4E79').replace('#', '')
    body_bg_color = header_config.get('tableBodyBgColor', '#FFFFFF').replace('#', '')
    
    for row in range(current_row, current_row + len(summary_data)):
        for col in range(1, len(summary_data[0]) + 1):
            cell = worksheet.cell(row=row, column=col)
            if row == current_row:  # Header row
                # Remove # if present for Excel compatibility
                header_color = header_bg_color
                if header_color.startswith('#'):
                    header_color = header_color[1:]
                cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            else:
                # Remove # if present for Excel compatibility
                body_color = body_bg_color
                if body_color.startswith('#'):
                    body_color = body_color[1:]
                cell.fill = PatternFill(start_color=body_color, end_color=body_color, fill_type='solid')
    
    pass

def auto_fit_excel_columns(worksheet) -> None:
    """Auto-fit column widths in Excel worksheet"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def create_pdf_header_elements(header_config: Dict[str, Any]) -> list:
    """Create PDF header elements from configuration"""
    elements = []
    styles = getSampleStyleSheet()
    
    # Create custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor(header_config.get('fontColor', '#1F4E79')),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor(header_config.get('fontColor', '#1F4E79')),
        alignment=TA_CENTER,
        spaceAfter=6
    )
    
    bank_style = ParagraphStyle(
        'BankInfo',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor(header_config.get('fontColor', '#1F4E79')),
        alignment=TA_CENTER,
        spaceAfter=3
    )
    
    # Add title and subtitle
    if header_config.get('includeHeader', True):
        elements.append(Paragraph(header_config.get('title', 'Dashboard Report'), title_style))
        elements.append(Paragraph(header_config.get('subtitle', 'Comprehensive Analysis Report'), subtitle_style))
        
        # Add bank information
        if header_config.get('showBankInfo', True):
            bank_name = header_config.get('bankName', DEFAULT_BANK_CONFIG['name'])
            bank_address = header_config.get('bankAddress', DEFAULT_BANK_CONFIG['address'])
            bank_phone = header_config.get('bankPhone', DEFAULT_BANK_CONFIG['phone'])
            bank_website = header_config.get('bankWebsite', DEFAULT_BANK_CONFIG['website'])
            
            elements.append(Paragraph(f"{DEFAULT_BANK_CONFIG['logo']} {bank_name}", bank_style))
            elements.append(Paragraph(bank_address, bank_style))
            elements.append(Paragraph(f"Tel: {bank_phone} | Web: {bank_website}", bank_style))
        
        # Add generation date
        if header_config.get('showDate', True):
            current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            elements.append(Paragraph(f"Generated on: {current_date}", bank_style))
        
        elements.append(Spacer(1, 20))
    
    return elements

def create_pdf_footer_elements(header_config: Dict[str, Any]) -> list:
    """Create PDF footer elements from configuration"""
    elements = []
    styles = getSampleStyleSheet()
    
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    if header_config.get('footerShowDate', True):
        current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elements.append(Paragraph(f"Generated on: {current_date}", footer_style))
    
    if header_config.get('footerShowConfidentiality', True):
        confidentiality_text = header_config.get('footerConfidentialityText', 'Confidential Report - Internal Use Only')
        elements.append(Paragraph(confidentiality_text, footer_style))
    
    return elements
